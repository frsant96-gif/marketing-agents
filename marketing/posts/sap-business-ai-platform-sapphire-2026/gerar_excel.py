import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Posts"

headers = [
    "Data", "Dia", "Linha Editorial", "Tema", "Origem do Post", "Título de Conteúdo",
    "Objetivo", "Persona", "Etapa do Funil", "Copy LinkedIn", "Copy Facebook",
    "Copy Instagram", "Formato", "Horário", "Ref. Texto", "Responsável", "Status"
]
ws.append(headers)

linkedin_v1 = """A SAP não lançou uma ferramenta nova no SAPPHIRE 2026.

Ela reescreveu a arquitetura do ERP.

O nome é SAP Business AI Platform: uma fundação única que integra SAP BTP, SAP Business Data Cloud e SAP AI Foundation. É sobre ela que os agentes Joule executam processos de negócio de ponta a ponta.

A peça que faz isso funcionar não é o agente. É o Knowledge Graph — 50 anos de lógica de ERP codificados em relações que a IA consegue interpretar.

A SAP também confirmou a Anthropic como parceira de modelos de fundação — o Claude já está disponível no SAP AI Foundation para agentes Joule em RH, compras e supply chain.

Pra quem lidera dados ou tecnologia: a pergunta não é se vai adotar agentes de IA no SAP. É se a fundação de dados já está pronta pra isso.

Quem tem SAP BDC estruturado está no caminho certo. Quem opera com dados fragmentados vai acelerar erro, não resultado.

Onde sua empresa está nessa régua?

#SAPBusinessDataCloud #SAPBDC #SAPSapphire2026 #DataGovernance #ERP"""

instagram = """A SAP acabou de reescrever a arquitetura do ERP 🧠

No SAPPHIRE 2026, ela uniu BTP + SAP Business Data Cloud + AI Foundation em uma única plataforma: o SAP Business AI Platform.

O que isso muda na prática:
→ Os agentes Joule agora executam processos completos, não só respondem perguntas
→ Eles decidem com base num Knowledge Graph que carrega 50 anos de lógica de ERP
→ A Anthropic entrou como parceira — o Claude já roda dentro do SAP AI Foundation

Agente de IA sobre dado fragmentado não acelera resultado. Acelera erro.

📌 Salva esse post se você lidera dados ou tecnologia na sua empresa.

#SAPBusinessDataCloud #SAPBDC #InteligenciaArtificial #DadosCorporativos #SAPSapphire #ERP #Analytics #GestaoDeDados #TransformacaoDeDados #CIO"""

facebook = """A SAP reescreveu a arquitetura do ERP no SAPPHIRE 2026.

O SAP Business AI Platform integra SAP BTP, SAP Business Data Cloud e SAP AI Foundation em uma fundação única — é sobre ela que os agentes Joule executam processos de ponta a ponta, com apoio de um Knowledge Graph que codifica 50 anos de lógica de ERP.

A Anthropic entrou como parceira de modelos: o Claude já roda no SAP AI Foundation.

Quem já tem SAP BDC estruturado está na base certa pra dar o próximo passo. Quem opera com dados fragmentados precisa resolver essa fundação antes de pensar em agentes."""

row = [
    "", "",
    "Autoridade/Tendência",
    "SAP Business AI Platform e o que muda pra quem já tem SAP BDC",
    "Artigo do blog",
    "A SAP não lançou uma ferramenta nova no SAPPHIRE 2026. Ela reescreveu a arquitetura do ERP.",
    "Educação",
    "CIO / Head de Dados",
    "Topo",
    linkedin_v1,
    facebook,
    instagram,
    "Texto longo",
    "",
    "marketing/posts/sap-business-ai-platform-sapphire-2026/copy.md",
    "Fran",
    "🟡 Criado",
]
ws.append(row)

header_fill = PatternFill(start_color="0A1A3C", end_color="0A1A3C", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)
alt_fill = PatternFill(start_color="EEF2F8", end_color="EEF2F8", fill_type="solid")
status_yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
status_yellow_font = Font(color="7A5A00")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="top", wrap_text=True)

for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r % 2 == 0:
            cell.fill = alt_fill

status_col = headers.index("Status") + 1
for r in range(2, ws.max_row + 1):
    cell = ws.cell(row=r, column=status_col)
    if cell.value == "🟡 Criado":
        cell.fill = status_yellow_fill
        cell.font = status_yellow_font

copy_cols = ["Copy LinkedIn", "Copy Facebook", "Copy Instagram"]
widths = {
    "Data": 12, "Dia": 10, "Linha Editorial": 18, "Tema": 30, "Origem do Post": 16,
    "Título de Conteúdo": 40, "Objetivo": 14, "Persona": 18, "Etapa do Funil": 14,
    "Copy LinkedIn": 68, "Copy Facebook": 68, "Copy Instagram": 68,
    "Formato": 14, "Horário": 10, "Ref. Texto": 45, "Responsável": 12, "Status": 12,
}
for i, h in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)

ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

wb.save("post-dados.xlsx")
print("ok")

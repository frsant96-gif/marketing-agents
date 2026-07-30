import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Posts"

headers = [
    "Data", "Dia", "Linha Editorial", "Tema", "Origem do Post", "Título de Conteúdo",
    "Objetivo", "Persona", "Etapa do Funil", "Copy LinkedIn", "Copy Facebook",
    "Copy Instagram", "Formato", "Horário", "Ref. Texto", "Responsável", "Status"
]
ws.append(headers)

linkedin_v1 = """A maioria das empresas trata o go-live do ERP como linha de chegada.

É o erro mais comum — e o mais caro.

Segundo o Panorama Mercado de Software (Portal ERP), 33,31% das empresas pretendem adquirir ou substituir seu ERP até 2026. Quem já implementou sabe: o sistema funcionar não significa que ele está gerando valor.

É aí que entra a sustentação analítica — o conjunto de processos e especialistas que garante que a camada de dados do ERP continue íntegra depois do go-live. Ela atua em 3 frentes:

→ Única fonte da verdade — o dado sai certo do ERP e chega certo no Power BI ou SAP Analytics Cloud
→ Governança e qualidade — identifica erros de cadastro antes que virem decisão errada
→ Evolução de contexto — atualiza regra de negócio conforme o mercado muda

Implementar ERP sem sustentação analítica é como comprar um avião de última geração sem equipe de manutenção.

Sua empresa passou por um go-live recente? A pergunta certa não é "o sistema está de pé". É "o sistema está gerando decisão".

#SustentaçãoAnalítica #ERP #GestãoDeDados #GoLive #BusinessIntelligence"""

instagram = """Seu ERP passou pelo go-live. E agora? 📊

A maioria das empresas trata o go-live como linha de chegada.
Mas é só o começo de uma nova etapa.

O que garante que o ERP continue gerando valor:
→ Dado sai certo do sistema e chega certo no Power BI ou SAP Analytics Cloud
→ Erros de cadastro são identificados antes de virarem decisão errada
→ Regras de negócio são atualizadas conforme o mercado muda

Isso se chama sustentação analítica.

Sem ela, implementar ERP é como comprar um avião de última geração sem equipe de manutenção ✈️

33,31% das empresas vão trocar ou adquirir um ERP até 2026 (Panorama Mercado de Software). Quem já implementou não pode repetir o erro de achar que o trabalho terminou no go-live.

📌 Salva esse post se sua empresa passou por um go-live recente.

#SustentaçãoAnalítica #ERP #GestãoDeDados #PowerBI #SAPAnalyticsCloud #BusinessIntelligence #GoLive #Analytics #DataGovernance #ConsultoriaEmpresarial"""

facebook = """A maioria das empresas trata o go-live do ERP como linha de chegada — e esse é o erro mais caro.

Segundo o Panorama Mercado de Software (Portal ERP), 33,31% das empresas vão adquirir ou substituir seu ERP até 2026.

A sustentação analítica garante que o dado saia certo do ERP e chegue certo no Power BI ou SAP Analytics Cloud, com governança e regras de negócio sempre atualizadas.

Implementar ERP sem sustentação analítica é como comprar um avião de última geração sem equipe de manutenção. A Solveplan pode ajudar."""

row = [
    "", "",
    "Educação/Produto",
    "Sustentação analítica pós go-live de ERP",
    "Artigo do blog",
    "A maioria das empresas trata o go-live do ERP como linha de chegada. É o erro mais comum — e o mais caro.",
    "Educação",
    "CIO / Head de Dados",
    "Topo",
    linkedin_v1,
    facebook,
    instagram,
    "Texto longo",
    "",
    "marketing/posts/sustentacao-analitica-pos-golive-erp/copy.md",
    "Fran",
    "🟡 Criado",
]
ws.append(row)

header_fill = PatternFill(start_color="0A1A3C", end_color="0A1A3C", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)
alt_fill = PatternFill(start_color="EEF2F8", end_color="EEF2F8", fill_type="solid")
status_yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
status_yellow_font = Font(color="7A5A00")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="top", wrap_text=True)

for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r % 2 == 0:
            cell.fill = alt_fill

status_col = headers.index("Status") + 1
for r in range(2, ws.max_row + 1):
    cell = ws.cell(row=r, column=status_col)
    if cell.value == "🟡 Criado":
        cell.fill = status_yellow_fill
        cell.font = status_yellow_font

widths = {
    "Data": 12, "Dia": 10, "Linha Editorial": 18, "Tema": 30, "Origem do Post": 16,
    "Título de Conteúdo": 40, "Objetivo": 14, "Persona": 18, "Etapa do Funil": 14,
    "Copy LinkedIn": 68, "Copy Facebook": 68, "Copy Instagram": 68,
    "Formato": 14, "Horário": 10, "Ref. Texto": 45, "Responsável": 12, "Status": 12,
}
for i, h in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)

ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

wb.save("post-dados.xlsx")
print("ok")

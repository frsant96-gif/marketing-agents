"""Atualiza PowerBI_Solveplan_H1_2026.xlsx com dados ao vivo do HubSpot (snapshot 25/07/2026).
Adiciona 2 abas novas (fReceita_Mensal, fPipeline_Etapas_Jul26) e novas linhas na KPIs_Resumo,
sem sobrescrever os dados originais Jan-Mai (leads/eventos/site/ads)."""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = "PowerBI_Solveplan_H1_2026.xlsx"

wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color="006AFF", end_color="006AFF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_table(ws, headers, rows, start_row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for r, row in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    for c, h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + c)].width = max(14, len(str(h)) + 2)


# --- Nova aba: receita mensal fechada (Ganho), por mes de fechamento, Pipeline de Vendas ---
if "fReceita_Mensal" in wb.sheetnames:
    del wb["fReceita_Mensal"]
ws1 = wb.create_sheet("fReceita_Mensal")
receita_rows = [
    ("Jan", 1, 2026, 2, 2494180.00, "Jan-Jul 2026 (HubSpot live 25/07)"),
    ("Fev", 2, 2026, 3, 4978346.74, "Jan-Jul 2026 (HubSpot live 25/07)"),
    ("Mar", 3, 2026, 1, 150000.00, "Jan-Jul 2026 (HubSpot live 25/07)"),
    ("Abr", 4, 2026, 2, 303583.68, "Jan-Jul 2026 (HubSpot live 25/07)"),
    ("Mai", 5, 2026, 6, 2606158.00, "Jan-Jul 2026 (HubSpot live 25/07)"),
    ("Jun", 6, 2026, 0, 0.00, "Jan-Jul 2026 (HubSpot live 25/07)"),
    ("Jul", 7, 2026, 5, 936584.00, "Jan-Jul 2026 (HubSpot live 25/07, parcial ate dia 25)"),
]
write_table(
    ws1,
    ["Mes_Nome", "Mes_Num", "Ano", "Negocios_Ganhos", "Receita_BRL", "Periodo"],
    receita_rows,
)

# --- Nova aba: pipeline aberto por etapa (snapshot 25/07/2026), Pipeline de Vendas ---
if "fPipeline_Etapas_Jul26" in wb.sheetnames:
    del wb["fPipeline_Etapas_Jul26"]
ws2 = wb.create_sheet("fPipeline_Etapas_Jul26")
pipeline_rows = [
    ("Conexao / Alinhamento Inicial", 3, 3800000.00, "Ativo"),
    ("Discovery / Escopo", 12, 3970500.00, "Ativo"),
    ("Proposta Tecnica e Escopo", 5, 3698676.52, "Ativo"),
    ("Aprovacao Precificacao", 1, 2400000.00, "Ativo"),
    ("Proposta enviada", 12, 7112420.12, "Ativo"),
    ("Negociacao", 5, 3587000.00, "Ativo"),
    ("Juridico / Compliance", 2, 810000.00, "Ativo"),
    ("Assinatura", 2, 726800.00, "Ativo"),
    ("Sleep Deals", 20, 7218000.00, "Ativo (parado / risco)"),
]
write_table(
    ws2,
    ["Etapa", "Qtd_Negocios", "Valor_BRL", "Status"],
    pipeline_rows,
)
ws2.cell(row=len(pipeline_rows) + 3, column=1, value="Total").font = Font(bold=True)
ws2.cell(row=len(pipeline_rows) + 3, column=2, value=sum(r[1] for r in pipeline_rows)).font = Font(bold=True)
ws2.cell(row=len(pipeline_rows) + 3, column=3, value=sum(r[2] for r in pipeline_rows)).font = Font(bold=True)
ws2.cell(row=len(pipeline_rows) + 4, column=1, value="Fonte: HubSpot MCP, query ao vivo em 25/07/2026, Pipeline de Vendas (default)")

# --- Atualiza KPIs_Resumo: adiciona linhas novas (nao sobrescreve as Jan-Mai) ---
ws_kpi = wb["KPIs_Resumo"]
next_row = ws_kpi.max_row + 1
novos_kpis = [
    ("Negocios Ganhos (Jan-Jul, fechamento)", 19, None, "OK", "Jan-Jul 2026 (HubSpot live)"),
    ("Receita Gerada Jan-Jul (R$)", 11469852.42, None, "OK", "Jan-Jul 2026 (HubSpot live)"),
    ("Pipeline Aberto atual (R$)", 33323396.64, None, "OK", "Snapshot 25/07/2026"),
    ("Negocios Ativos no Pipeline", 62, None, "OK", "Snapshot 25/07/2026"),
    ("Negocios Perdidos (Jan-Jul, criacao)", 24, None, "ATENCAO", "Jan-Jul 2026 (HubSpot live)"),
    ("Valor Perdido (R$)", 4690000, None, "ATENCAO", "Jan-Jul 2026 (HubSpot live)"),
    ("Taxa de Fechamento (Ganho vs Perdido)", 44.2, None, "ATENCAO", "Jan-Jul 2026 (HubSpot live)"),
    ("Negocios em Sleep Deals (parados)", 20, None, "ATENCAO", "Snapshot 25/07/2026 - R$ 7,22M em risco"),
]
for i, row in enumerate(novos_kpis):
    for c, val in enumerate(row, start=1):
        ws_kpi.cell(row=next_row + i, column=c, value=val)

wb.save(PATH)
print("Workbook atualizado:", PATH)
print("Abas:", wb.sheetnames)

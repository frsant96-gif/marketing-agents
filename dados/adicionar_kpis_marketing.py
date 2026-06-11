import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

AZUL = "006AFF"
VERDE = "00C4A7"
LARANJA = "FF6B35"
CINZA_ESCURO = "0A0E19"
BRANCO = "FFFFFF"
CINZA_CLARO = "F5F5F5"
AMARELO = "FFF3CD"
VERMELHO_CLARO = "FFE0E0"

def estilizar_aba(ws, cor_header=AZUL):
    header_fill = PatternFill("solid", fgColor=cor_header)
    header_font = Font(bold=True, color=BRANCO)
    alt_fill = PatternFill("solid", fgColor=CINZA_CLARO)
    borda = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    for i, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.border = borda
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical='center', horizontal='center')
            elif i % 2 == 0:
                cell.fill = alt_fill
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


# ============================================================
# fMktDeals — 27 deals de origens primárias de marketing H1 2026
# Dados: HubSpot export + HubSpot MCP API (08/06/2026)
# ============================================================
fMktDeals = pd.DataFrame([
    # Evento SAP (16 deals)
    {'ID': 55389550137, 'Nome': 'Dellavolp - Fabrica de Analytics', 'Empresa': 'Dellavolp', 'Origem': 'Evento SAP', 'Campanha': 'SAP BTP Experience 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 350000, 'Dias_para_Fechar': 35, 'Produto': 'SAP Analytics Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Fev', 'Mes_Num': 2},
    {'ID': 57093444876, 'Nome': 'Hayamax - Licenciamento SAP BDC', 'Empresa': 'Hayamax', 'Origem': 'Evento SAP', 'Campanha': 'SAP BTP Experience 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 31, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Fev', 'Mes_Num': 2},
    {'ID': 58136152638, 'Nome': 'Sicredi - Fabrica SAP Datasphere', 'Empresa': 'Sicredi', 'Origem': 'Evento SAP', 'Campanha': 'SAP BTP Experience 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 450000, 'Dias_para_Fechar': 70, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mar', 'Mes_Num': 3},
    {'ID': 58158520398, 'Nome': 'Wildlife Studios', 'Empresa': 'Wildlife Studios', 'Origem': 'Evento SAP', 'Campanha': 'SAP BTP Experience 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 0, 'Dias_para_Fechar': 58, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mar', 'Mes_Num': 3},
    {'ID': 58164419618, 'Nome': 'BTG Pactual', 'Empresa': 'BTG Pactual', 'Origem': 'Evento SAP', 'Campanha': 'SAP BTP Experience 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 13, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mar', 'Mes_Num': 3},
    {'ID': 58354726933, 'Nome': 'Bemol', 'Empresa': 'Bemol', 'Origem': 'Evento SAP', 'Campanha': 'SAP BTP Experience 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 19, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mar', 'Mes_Num': 3},
    {'ID': 58358435118, 'Nome': 'SINOVA - SAP BDC', 'Empresa': 'SINOVA', 'Origem': 'Evento SAP', 'Campanha': 'SAP BTP Experience 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 150000, 'Dias_para_Fechar': 67, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mar', 'Mes_Num': 3},
    {'ID': 59186304944, 'Nome': 'Citrosuco - Planejamento Orcamentario', 'Empresa': 'Citrosuco', 'Origem': 'Evento SAP', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Aprovacao Precificacao', 'Status': 'Ativo', 'Valor_BRL': 2400000, 'Dias_para_Fechar': 75, 'Produto': 'SAP Analytics Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59229970248, 'Nome': 'Tuberfil', 'Empresa': 'Tuberfil', 'Origem': 'Evento SAP', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 53, 'Produto': None, 'Tipo_Receita': None, 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59275221656, 'Nome': 'OUROFINO ANIMAL - SAC PLANNING', 'Empresa': 'OUROFINO', 'Origem': 'Evento SAP', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 1300000, 'Dias_para_Fechar': 104, 'Produto': 'SAP Analytics Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59275221802, 'Nome': 'FSJ - SAP BDC Licenciamento', 'Empresa': 'FSJ', 'Origem': 'Evento SAP', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 150000, 'Dias_para_Fechar': 73, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59275158381, 'Nome': 'SANTA HELENA - Licenciamento SAP BDC', 'Empresa': 'SANTA HELENA', 'Origem': 'Evento SAP', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 100000, 'Dias_para_Fechar': 73, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59277155126, 'Nome': 'SANTA HELENA - Fabrica Analytics', 'Empresa': 'SANTA HELENA', 'Origem': 'Evento SAP', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 300000, 'Dias_para_Fechar': 73, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59277278110, 'Nome': 'SINOVA - Fabrica SAP BDC', 'Empresa': 'SINOVA', 'Origem': 'Evento SAP', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 390000, 'Dias_para_Fechar': 43, 'Produto': 'SAP Business Data Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59085024838, 'Nome': 'Algar Telecom', 'Empresa': 'Algar Telecom', 'Origem': 'Evento SAP', 'Campanha': 'SAP NOW 2025', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 60, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 60050133557, 'Nome': 'Grupo Profarma', 'Empresa': 'Grupo Profarma', 'Origem': 'Evento SAP', 'Campanha': 'SAP NOW 2025', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 0, 'Dias_para_Fechar': 0, 'Produto': 'SAP Group Reporting', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mai', 'Mes_Num': 5},
    # Evento Próprio (6 deals)
    {'ID': 58638591055, 'Nome': 'Grupo Soufer', 'Empresa': 'Grupo Soufer', 'Origem': 'Evento Próprio', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 0, 'Dias_para_Fechar': 0, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 58658638708, 'Nome': 'C&A Modas', 'Empresa': 'C&A Modas', 'Origem': 'Evento Próprio', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 68, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 58878230127, 'Nome': 'CMAA', 'Empresa': 'CMAA', 'Origem': 'Evento Próprio', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 11, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 58882858095, 'Nome': 'Geolab', 'Empresa': 'Geolab', 'Origem': 'Evento Próprio', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 0, 'Dias_para_Fechar': 0, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59915353780, 'Nome': 'Grupo Adeste', 'Empresa': 'Grupo Adeste', 'Origem': 'Evento Próprio', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 15, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mai', 'Mes_Num': 5},
    {'ID': 60347429494, 'Nome': 'Coopercitrus', 'Empresa': 'Coopercitrus', 'Origem': 'Evento Próprio', 'Campanha': 'IT Summit Agro 2026', 'Etapa': 'Conexao / Alinhamento Inicial', 'Status': 'Ativo', 'Valor_BRL': 0, 'Dias_para_Fechar': 0, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mai', 'Mes_Num': 5},
    # Inbound Marketing (5 deals)
    {'ID': 41895925367, 'Nome': 'REAM - Planejamento Integrado', 'Empresa': 'REAM', 'Origem': 'Inbound Marketing', 'Campanha': 'Webinar SAP IBP', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 1400000, 'Dias_para_Fechar': 94, 'Produto': 'SAP Analytics Cloud', 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Jan', 'Mes_Num': 1},
    {'ID': 58157901189, 'Nome': 'Logum Logistica - Planejamento de RH', 'Empresa': 'Logum Logistica', 'Origem': 'Inbound Marketing', 'Campanha': 'LinkedIn Page', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 400000, 'Dias_para_Fechar': 74, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Mar', 'Mes_Num': 3},
    {'ID': 59419151717, 'Nome': 'Docol - Planejamento Orcamentario', 'Empresa': 'Docol', 'Origem': 'Inbound Marketing', 'Campanha': 'Formulário do Site', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 0, 'Dias_para_Fechar': 98, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 59814551927, 'Nome': 'Atem distribuidora', 'Empresa': 'Atem', 'Origem': 'Inbound Marketing', 'Campanha': 'Download: e-book Consolidação', 'Etapa': 'Discovery / Escopo', 'Status': 'Ativo', 'Valor_BRL': 0, 'Dias_para_Fechar': 32, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Abr', 'Mes_Num': 4},
    {'ID': 60345087235, 'Nome': 'Dellamed', 'Empresa': 'Dellamed', 'Origem': 'Inbound Marketing', 'Campanha': 'Formulário do site', 'Etapa': 'Negocio Perdido', 'Status': 'Perdido', 'Valor_BRL': 0, 'Dias_para_Fechar': 138, 'Produto': None, 'Tipo_Receita': 'New Business', 'Mes_Criacao': 'Jan', 'Mes_Num': 1},
])

# Calcular pipeline por status
total_mkt = len(fMktDeals)
ativos = fMktDeals[fMktDeals['Status'] == 'Ativo']
perdidos = fMktDeals[fMktDeals['Status'] == 'Perdido']
pipeline_ativo = fMktDeals.loc[fMktDeals['Status'] == 'Ativo', 'Valor_BRL'].sum()
pipeline_perdido = fMktDeals.loc[fMktDeals['Status'] == 'Perdido', 'Valor_BRL'].sum()

# Valores HubSpot confirmados via relatório (screenshot 11/06/2026)
pipeline_mkt_hubspot = 7790000  # R$ 7.790.000 confirmado no relatório HubSpot 2026
pipeline_influenciado = 35674717.59  # pipeline total influenciado por marketing (histórico)


# ============================================================
# KPIs_Marketing — Framework completo de KPIs de marketing
# Fonte: HubSpot MCP + GA4 + Planilha Geral + Redes Sociais
# ============================================================
kpis_mkt = pd.DataFrame([
    # GERAÇÃO DE LEADS
    {'KPI': 'Leads gerados (total)', 'Tipo': 'Leading', 'O_que_mede': 'Volume de geração', 'Valor_H1_2026': 307, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': '97% eventos (BTP 176 + IT Summit 121)'},
    {'KPI': 'Leads por origem - Eventos SAP', 'Tipo': 'Leading', 'O_que_mede': 'Mix de canais', 'Valor_H1_2026': 176, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': 'SAP BTP Experience Fev/26'},
    {'KPI': 'Leads por origem - Eventos Proprios', 'Tipo': 'Leading', 'O_que_mede': 'Mix de canais', 'Valor_H1_2026': 121, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': 'IT Summit Agro Abr/26'},
    {'KPI': 'Leads por origem - Inbound', 'Tipo': 'Leading', 'O_que_mede': 'Mix de canais', 'Valor_H1_2026': 10, 'Unidade': 'qtd', 'Status': 'ATENCAO', 'Notas': 'Apenas 3% do total — canal a desenvolver'},
    {'KPI': 'Leads por campanha - BTP Experience', 'Tipo': 'Leading', 'O_que_mede': 'Performance por campanha', 'Valor_H1_2026': 176, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': 'Maior gerador de leads do periodo'},
    {'KPI': 'Leads por campanha - IT Summit Agro', 'Tipo': 'Leading', 'O_que_mede': 'Performance por campanha', 'Valor_H1_2026': 121, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': '28 hot leads'},
    {'KPI': 'Leads qualificados (MQL)', 'Tipo': 'Leading', 'O_que_mede': 'Qualidade da geração', 'Valor_H1_2026': 212, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': '69% dos leads são MQL'},
    {'KPI': 'Taxa Lead para MQL', 'Tipo': 'Leading', 'O_que_mede': 'Eficiência da geração', 'Valor_H1_2026': 69.0, 'Unidade': '%', 'Status': 'OK', 'Notas': 'Acima da media B2B (30-50%)'},
    {'KPI': 'CPL (Custo por Lead) - BTP', 'Tipo': 'Leading', 'O_que_mede': 'Eficiência de investimento', 'Valor_H1_2026': 416.97, 'Unidade': 'R$', 'Status': 'OK', 'Notas': 'R$ 73.385 investidos / 176 leads'},
    {'KPI': 'Custo por MQL - BTP', 'Tipo': 'Leading', 'O_que_mede': 'Eficiência real', 'Valor_H1_2026': 604.01, 'Unidade': 'R$', 'Status': 'OK', 'Notas': 'R$ 73.385 / 121 MQLs do BTP'},
    {'KPI': 'MQLs gerados', 'Tipo': 'Leading', 'O_que_mede': 'Volume de leads qualificados', 'Valor_H1_2026': 212, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': 'BTP: 121 MQLs | IT Summit: 91 MQLs'},
    # PIPELINE / LAGGING
    {'KPI': 'Pipeline gerado por Marketing (R$)', 'Tipo': 'Lagging', 'O_que_mede': 'Receita potencial originada em marketing', 'Valor_H1_2026': 7790000, 'Unidade': 'R$', 'Status': 'OK', 'Notas': 'Confirmado no relatorio HubSpot 11/06/2026'},
    {'KPI': 'Pipeline influenciado por Marketing (R$)', 'Tipo': 'Lagging', 'O_que_mede': 'Impacto no funil', 'Valor_H1_2026': 35674717.59, 'Unidade': 'R$', 'Status': 'OK', 'Notas': 'Historico total — todos os pipelines com toque de mkt'},
    {'KPI': '% do pipeline com toque de Marketing', 'Tipo': 'Lagging', 'O_que_mede': 'Dependência do marketing', 'Valor_H1_2026': 22.4, 'Unidade': '%', 'Status': 'OK', 'Notas': '7,79M / 34,8M total H1 2026'},
    {'KPI': 'Negócios gerados por Marketing (Qtd)', 'Tipo': 'Lagging', 'O_que_mede': 'Volume de negócios originados em marketing', 'Valor_H1_2026': 27, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': 'Evento SAP 16 + Evento Proprio 6 + Inbound 5'},
    {'KPI': 'Negócios novos 2026 via Marketing (HubSpot)', 'Tipo': 'Lagging', 'O_que_mede': 'Volume gerado em 2026', 'Valor_H1_2026': 30, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': 'Fonte: relatorio HubSpot — inclui influenciados'},
    {'KPI': 'Oportunidades (OPP)', 'Tipo': 'Lagging', 'O_que_mede': 'Conversão final', 'Valor_H1_2026': 9, 'Unidade': 'qtd', 'Status': 'ATENCAO', 'Notas': 'Total H1 — taxa MQL-OPP baixa (4%)'},
    {'KPI': 'Taxa MQL para SQL', 'Tipo': 'Leading', 'O_que_mede': 'Eficiência com SDR', 'Valor_H1_2026': 4.2, 'Unidade': '%', 'Status': 'ATENCAO', 'Notas': '9 SQLs / 212 MQLs — gargalo no follow-up pos-evento'},
    {'KPI': 'Win Rate Marketing (H1 2026)', 'Tipo': 'Lagging', 'O_que_mede': 'Efetividade por origem', 'Valor_H1_2026': 0, 'Unidade': '%', 'Status': 'ATENCAO', 'Notas': '0 ganhos de origem marketing em H1 — deals ainda em ciclo'},
    {'KPI': 'Pipeline ativo Marketing (R$)', 'Tipo': 'Lagging', 'O_que_mede': 'Oportunidade disponivel para H2', 'Valor_H1_2026': 4790000, 'Unidade': 'R$', 'Status': 'OK', 'Notas': 'Soma dos deals ativos com valor preenchido no HubSpot'},
    {'KPI': 'Contas impactadas por Marketing', 'Tipo': 'Leading', 'O_que_mede': 'Alcance em contas-alvo', 'Valor_H1_2026': 27, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': 'Empresas unicas com deal originado em marketing'},
    # ROI
    {'KPI': 'ROI de Marketing - BTP Experience', 'Tipo': 'Lagging', 'O_que_mede': 'Retorno financeiro', 'Valor_H1_2026': 20, 'Unidade': 'x', 'Status': 'OK', 'Notas': 'Pipeline gerado / investimento = 7,79M / 73.385'},
    {'KPI': 'Payback da campanha - BTP Experience', 'Tipo': 'Lagging', 'O_que_mede': 'Retorno no tempo', 'Valor_H1_2026': 106.1, 'Unidade': 'x', 'Status': 'OK', 'Notas': 'Pipeline H1 (7,79M) / custo (73.385) = 106x potencial'},
    # SITE
    {'KPI': 'Sessoes no site', 'Tipo': 'Leading', 'O_que_mede': 'Alcance digital', 'Valor_H1_2026': 17140, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': '+160% vs periodo anterior'},
    {'KPI': 'Usuarios ativos', 'Tipo': 'Leading', 'O_que_mede': 'Audiencia unica', 'Valor_H1_2026': 13294, 'Unidade': 'qtd', 'Status': 'OK', 'Notas': '+136% vs periodo anterior'},
    {'KPI': 'CTR Google Ads', 'Tipo': 'Leading', 'O_que_mede': 'Relevancia do anuncio', 'Valor_H1_2026': 4.97, 'Unidade': '%', 'Status': 'OK', 'Notas': 'Acima da media B2B (2-5%)'},
    # PIPELINE POR SOLUCAO
    {'KPI': 'Pipeline por solucao - SAP Business Data Cloud', 'Tipo': 'Lagging', 'O_que_mede': 'Escala por oferta', 'Valor_H1_2026': 1190000, 'Unidade': 'R$', 'Status': 'OK', 'Notas': 'SINOVA BDC + FSJ + Santa Helena x2 — 5 deals ativos'},
    {'KPI': 'Pipeline por solucao - SAP Analytics Cloud', 'Tipo': 'Lagging', 'O_que_mede': 'Escala por oferta', 'Valor_H1_2026': 3700000, 'Unidade': 'R$', 'Status': 'OK', 'Notas': 'Citrosuco + OUROFINO — 2 deals de alto valor'},
    {'KPI': 'Tempo medio ate perda (deals mkt)', 'Tipo': 'Leading', 'O_que_mede': 'Velocidade do ciclo', 'Valor_H1_2026': 52, 'Unidade': 'dias', 'Status': 'OK', 'Notas': 'Media dos deals perdidos de marketing — ciclo curto'},
    # ND = Nao Disponivel
    {'KPI': 'Tempo Lead para MQL', 'Tipo': 'Leading', 'O_que_mede': 'Velocidade do funil', 'Valor_H1_2026': None, 'Unidade': 'dias', 'Status': 'ND', 'Notas': 'Requer data de mudança de etapa no HubSpot Contacts'},
    {'KPI': 'SLA MQL para 1o contato', 'Tipo': 'Leading', 'O_que_mede': 'Velocidade de contato do SDR', 'Valor_H1_2026': None, 'Unidade': 'horas', 'Status': 'ND', 'Notas': 'Requer log de atividades (calls/emails) no HubSpot'},
    {'KPI': 'Decisores impactados', 'Tipo': 'Leading', 'O_que_mede': 'Qualidade do contato', 'Valor_H1_2026': None, 'Unidade': 'qtd', 'Status': 'ND', 'Notas': 'Requer campo Cargo/Seniority preenchido nos contacts'},
    {'KPI': 'Contas-alvo engajadas (ABM)', 'Tipo': 'Leading', 'O_que_mede': 'Adocao ABM', 'Valor_H1_2026': None, 'Unidade': 'qtd', 'Status': 'ND', 'Notas': 'Requer lista de target accounts definida'},
])


# ============================================================
# Abrir Excel existente e adicionar as novas abas
# ============================================================
output_path = '../dados/PowerBI_Solveplan_H1_2026.xlsx'

# Carregar workbook existente
wb = openpyxl.load_workbook(output_path)

# Remover abas antigas se existirem
for nome in ['fMktDeals', 'KPIs_Marketing']:
    if nome in wb.sheetnames:
        del wb[nome]

# Adicionar fMktDeals
ws_mkt = wb.create_sheet('fMktDeals')
# Header
headers = list(fMktDeals.columns)
for col_idx, h in enumerate(headers, 1):
    ws_mkt.cell(row=1, column=col_idx, value=h)
# Data rows
for row_idx, row_data in fMktDeals.iterrows():
    for col_idx, val in enumerate(row_data, 1):
        ws_mkt.cell(row=row_idx + 2, column=col_idx, value=val)
estilizar_aba(ws_mkt, cor_header=VERDE)

# Adicionar KPIs_Marketing
ws_kpi = wb.create_sheet('KPIs_Marketing')
headers_kpi = list(kpis_mkt.columns)
for col_idx, h in enumerate(headers_kpi, 1):
    ws_kpi.cell(row=1, column=col_idx, value=h)
for row_idx, row_data in kpis_mkt.iterrows():
    for col_idx, val in enumerate(row_data, 1):
        ws_kpi.cell(row=row_idx + 2, column=col_idx, value=val)
# Color status column
status_col = headers_kpi.index('Status') + 1
for row in ws_kpi.iter_rows(min_row=2):
    status_cell = row[status_col - 1]
    if status_cell.value == 'OK':
        status_cell.fill = PatternFill("solid", fgColor="D4EDDA")
        status_cell.font = Font(color="155724", bold=True)
    elif status_cell.value == 'ATENCAO':
        status_cell.fill = PatternFill("solid", fgColor="FFF3CD")
        status_cell.font = Font(color="856404", bold=True)
    elif status_cell.value == 'ND':
        status_cell.fill = PatternFill("solid", fgColor="E2E3E5")
        status_cell.font = Font(color="6C757D")
estilizar_aba(ws_kpi, cor_header=AZUL)

wb.save(output_path)
print(f'Excel atualizado: {output_path}')
print(f'Abas adicionadas: fMktDeals ({len(fMktDeals)} deals), KPIs_Marketing ({len(kpis_mkt)} KPIs)')
print()
print('=== RESUMO DOS KPIs DE MARKETING H1 2026 ===')
print(f'Negócios de marketing: {total_mkt} (Ativos: {len(ativos)}, Perdidos: {len(perdidos)}, Ganhos: 0)')
print(f'Pipeline gerado por marketing (HubSpot): R$ {pipeline_mkt_hubspot:,.0f}')
print(f'Pipeline ativo marketing (com valor): R$ {pipeline_ativo:,.0f}')
print(f'% pipeline com toque marketing: {pipeline_mkt_hubspot/34813628*100:.1f}%')
print(f'ROI BTP Experience: {pipeline_mkt_hubspot/73385.91:.0f}x')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cronograma Junho 2026"

# ── Cabeçalhos ──────────────────────────────────────────────────────────────
headers = [
    "Data", "Dia", "Linha Editorial", "Tema", "Origem do Post",
    "Título de Conteúdo", "Objetivo", "Persona", "Etapa do Funil",
    "Copy LinkedIn", "Copy Facebook", "Copy Instagram",
    "Formato", "Horário", "Ref. Texto", "Responsável", "Status"
]

HEADER_BG   = "0A1A3C"   # azul escuro Solveplan
HEADER_FONT = "FFFFFF"
ROW_ALT     = "EEF2F8"   # azul bem claro para linhas pares
PRONTO_BG   = "D6F5E8"   # verde claro
CRIADO_BG   = "FFF9C4"   # amarelo claro
GREEN_FG    = "1A7A4A"
ORANGE_FG   = "7A5A00"
BORDER_COLOR = "CCCCCC"

thin = Side(style="thin", color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# cabeçalho
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = Font(bold=True, color=HEADER_FONT, size=10)
    cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border

ws.row_dimensions[1].height = 32

# ── Dados ────────────────────────────────────────────────────────────────────
posts = [
    # (Data, Dia, Linha Editorial, Tema, Origem, Título, Objetivo, Persona, Funil,
    #  Copy LI, Copy FB, Copy IG, Formato, Horário, Ref Texto, Responsável, Status)
    (
        "03/06 (Qua)", "Quarta",
        "Case / Prova Social",
        "Case Lins Agroindustrial — SAP Analytics Cloud",
        "Case de cliente",
        "De 8h para minutos: como a Lins Agroindustrial transformou o planejamento com SAP Analytics Cloud",
        "Credibilidade / Geração de pipeline",
        "CFO, Controller, Head de Dados — Agro",
        "Meio",
        "A @Lins Agroindustrial processava dados de planejamento em planilhas isoladas e gastava cerca de 8 horas para consolidar informações entre as áreas agrícola, industrial e financeira. Com a migração do SAP BPC para o SAP Analytics Cloud, implementada pela @SolvePlan, esse tempo caiu 90%. Hoje, cada etapa de cálculo é concluída em minutos, as informações estão centralizadas em dashboards confiáveis e o planejamento integrado entre todas as áreas virou realidade. O resultado não é só velocidade. É a diferença entre reagir ao que já aconteceu e antecipar o que pode acontecer. Assista ao case completo → [link] [MARCAR: @Ueliton | @Priscila] #SAPAnalyticsCloud #Agronegócio #PlanejamentoFinanceiro #SAP #FPeA",
        "A @Lins Agroindustrial saiu de 8 horas de consolidação para minutos, após migrar do SAP BPC para o SAP Analytics Cloud com a @SolvePlan. 90% de redução no tempo de processamento. Em um setor onde commodity não espera. Assista ao case completo → [link] #SAPAnalyticsCloud #Agronegócio #FPeA",
        "De 8h para minutos. 🌾 A Lins Agroindustrial transformou o planejamento financeiro com SAP Analytics Cloud. 90% menos tempo de processamento. Case completo no link da bio. #SAPAnalyticsCloud #Agronegócio #FPeA #SAP",
        "Texto longo", "9h00",
        "marketing/posts/case-lins-agroindustrial/copy.md",
        "Fran", "✅ Pronto"
    ),
    (
        "04/06 (Qui)", "Quinta",
        "Autoridade / Tendência",
        "SAPPHIRE 2026 — Fim do ERP como se conhece",
        "Evento SAP SAPPHIRE 2026",
        "O SAPPHIRE 2026 não foi sobre IA. Foi sobre o fim do ERP como você conhece.",
        "Awareness / Autoridade executiva",
        "CIO, CFO, CEO, Diretor de Dados",
        "Topo",
        "O SAPPHIRE 2026 não foi sobre IA. Foi sobre o fim do ERP como a maioria das empresas ainda opera. 50+ assistentes especializados. 200+ agentes. Fechamento financeiro de semanas para dias. €100 milhões em investimento de parceiros. E um case concreto: a Vale saiu de 100+ planilhas para -50% no tempo de reporting. O que isso significa pra quem decide sobre dados e planejamento nas empresas → nos slides. #SAPSapphire2026 #SAPBusinessDataCloud #AutonomousEnterprise #FPandA #BusinessAI #Solveplan",
        "O SAPPHIRE 2026 não foi sobre IA. Foi sobre o fim do ERP como a maioria ainda opera. 50+ assistentes. 200+ agentes. €100M em parceiros. Case Vale: -50% no tempo de reporting. O que isso muda pra você → veja os slides. #SAPSapphire2026 #SAPBusinessDataCloud #Solveplan",
        "O SAPPHIRE 2026 mudou tudo. 🚀 50+ assistentes, 200+ agentes, €100M de investimento. A Vale saiu de 100+ planilhas para -50% no reporting. Desliza para entender o que muda. #SAPPHIRE2026 #SAP #BDC #IA",
        "Carrossel (8 slides)", "10h00",
        "marketing/posts/pos-sapphire-2026/copy.md",
        "Fran", "✅ Pronto"
    ),
    (
        "05/06 (Sex)", "Sexta",
        "Educação / Produto",
        "O que é SAP BDC na prática",
        "Criação própria",
        "Toda empresa SAP tem dado. A maioria não tem dado confiável.",
        "Educação / Awareness de produto",
        "CFO, Controller, Head de Dados, CIO",
        "Topo",
        "Toda empresa SAP tem dado. A maioria não tem dado confiável. Existe uma diferença que poucos nomeiam: ter dado e ter dado que serve para decidir. Um dado que chegou ontem — mas precisa ser decidido hoje — já chegou errado. O SAP Business Data Cloud existe para resolver exatamente isso. Na prática, o BDC faz três coisas que a maioria dos ambientes SAP ainda não tem: → Integra dados com semântica de negócio (não só técnica) → Governa: quem acessa o quê, com qual nível de confiança → Contextualiza: o dado pertence a um processo, uma entidade e uma decisão. Para empresas pós-S/4HANA, esse é o próximo passo real. Se está avaliando SAP BDC, fala comigo. #SAPBusinessDataCloud #Analytics #Dados #FP&A #SAP #Solveplan",
        "Toda empresa SAP tem dado. Poucas têm dado que serve pra decidir. SAP Business Data Cloud resolve isso com 3 pilares: integração semântica, governança e contextualização por processo. Próximo passo real pós-S/4HANA. Quer saber mais? → [link] #SAPBusinessDataCloud #Analytics #Dados #Solveplan",
        "Toda empresa SAP tem dado. Poucas têm dado confiável. 🎯 O SAP BDC muda isso com integração semântica, governança e contextualização. Saiba mais: link na bio. #SAPBusinessDataCloud #Dados #FP&A #SAP",
        "Texto longo", "8h30",
        "marketing/posts/junho-2026/semana-1.md",
        "Fran", "🟡 Criado"
    ),
    (
        "08/06 (Seg)", "Segunda",
        "Artigo / Deep-dive",
        "SAP Business AI Platform — reconfiguração do ERP",
        "Artigo do blog",
        "A SAP não lançou uma nova ferramenta no SAPPHIRE 2026. Ela reconfigurou a arquitetura do ERP.",
        "Educação / Autoridade",
        "CIO, Diretor de TI, Head de Dados",
        "Topo / Meio",
        "A SAP não lançou uma nova ferramenta no SAPPHIRE 2026. Ela reconfigurou a arquitetura do ERP. BTP, SAP BDC e AI Foundation agora operam como uma fundação única. É sobre ela que os agentes Joule executam processos de ponta a ponta. → Automação com dado fragmentado consolida erros mais rápido do que qualquer equipe humana consegue corrigir. → Quem já tem SAP BDC está no lugar certo. Quem não tem, chega a essa arquitetura com uma lacuna real. Escrevemos um artigo explicando o que mudou e o que fazer a partir de agora. → Link no primeiro comentário. #SAPBusinessAIPlatform #SAPPHIRE2026 #SAPBusinessDataCloud #IA #SAP #Solveplan",
        "A SAP não lançou nova ferramenta no SAPPHIRE 2026 — reconfigurou o ERP. BTP + BDC + AI Foundation = uma fundação única para agentes. Automação sobre dado fragmentado consolida erros mais rápido do que qualquer equipe corrige. Artigo completo → [link] #SAPBusinessAIPlatform #SAPBusinessDataCloud #IA",
        "SAP Business AI Platform: não é nova ferramenta. É nova arquitetura. 🔧 BTP + BDC + AI Foundation integrados. Artigo no link da bio. #SAPBusinessAIPlatform #SAPPHIRE2026 #SAP #IA",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-2.md | marketing/blog/sap-business-ai-platform/artigo.md",
        "Fran", "🟡 Criado"
    ),
    (
        "09/06 (Ter)", "Terça",
        "Educação / Produto",
        "SAP Datasphere vs SAP BDC — diferenças práticas",
        "Criação própria",
        "SAP Datasphere vs SAP Business Data Cloud — qual a diferença e quando usar cada um?",
        "Educação / Consideração",
        "CIO, Head de Dados, Arquiteto SAP",
        "Meio",
        "SAP Datasphere vs SAP Business Data Cloud — qual a diferença? Muita gente confunde os dois. A resposta curta: Datasphere é onde o dado existe. BDC é onde o dado serve para decidir. Em 5 slides: o que cada um faz, para quem é, e qual usar no seu momento. Qual a sua situação hoje? Comenta abaixo. #SAPDatasphere #SAPBusinessDataCloud #Analytics #Dados #SAP #Solveplan",
        "SAP Datasphere vs SAP BDC — você sabe a diferença? Datasphere: onde o dado existe. BDC: onde o dado serve pra decidir. 5 slides explicando quando usar cada um → [link] #SAPDatasphere #SAPBusinessDataCloud #Analytics",
        "Datasphere vs BDC — você sabe a diferença? 📊 5 slides que explicam de uma vez. Desliza! #SAPDatasphere #SAPBusinessDataCloud #SAP #Dados",
        "Carrossel (5 slides)", "10h00",
        "marketing/posts/junho-2026/semana-2.md",
        "Fran", "🟡 Criado"
    ),
    (
        "10/06 (Qua)", "Quarta",
        "Engagement",
        "Poll — maior gargalo de dados",
        "Criação própria",
        "Qual é o maior gargalo de dados na sua empresa hoje?",
        "Engajamento / Geração de dados de audiência",
        "CFO, Controller, Head de Dados, CIO",
        "Topo",
        "Qual é o maior gargalo de dados na sua empresa hoje? A) Planilhas Excel como sistema de consolidação B) Dados em silos — cada área tem a sua versão C) Dado existe mas não é confiável / há divergências D) Falta de governança — ninguém sabe qual versão é a certa. A resposta mais comum que ouvimos em diagnósticos é C ou D. Qual é o seu? Vote abaixo. #Dados #Analytics #FP&A #SAP #Governança #Solveplan",
        "Qual o maior gargalo de dados na sua empresa? A: Planilhas | B: Silos | C: Dado não confiável | D: Falta de governança. O mais comum que vemos: C ou D. Qual é o seu? 👇 #Dados #FP&A #SAP #Governança",
        "Qual o maior gargalo de dados da sua empresa? 🤔 A: Planilhas | B: Silos | C: Dado não confiável | D: Sem governança. Conta nos comentários! #Dados #FP&A #SAP",
        "Poll (enquete)", "12h00",
        "marketing/posts/junho-2026/semana-2.md",
        "Fran", "🟡 Criado"
    ),
    (
        "11/06 (Qui)", "Quinta",
        "Autoridade / Tendência",
        "SAP + Anthropic para FP&A",
        "Criação própria",
        "A SAP firmou parceria com a Anthropic. E a maioria das notícias perdeu o ponto mais importante.",
        "Awareness / Autoridade executiva",
        "CFO, Controller, Diretor Financeiro",
        "Topo / Meio",
        "A SAP firmou parceria com a Anthropic. E a maioria das notícias sobre isso perdeu o ponto mais importante. Não é uma parceria de chatbot. É uma parceria de processo financeiro. O que viabiliza para quem fecha o mês no SAP: → Análise de variação Real x Budget em linguagem natural → Fechamento financeiro com verificação automatizada de consistência → Projeção de caixa com sugestão de ação corretiva. O que diferencia: o Claude opera sobre dados reais da empresa, governados pelo SAP BDC, dentro das políticas de aprovação configuradas. A pergunta não é se vão adotar IA. É sobre qual fundação de dados vão construir. #SAPAnthropic #FP&A #IA #SAP #ClaudeAI #Solveplan",
        "SAP + Anthropic não é parceria de chatbot. É parceria de processo financeiro. Análise de variação Real x Budget em linguagem natural. Fechamento com verificação automática. Projeção de caixa com ação corretiva. Tudo com dado real do SAP, não da internet. #SAPAnthropic #FP&A #IA #SAP",
        "SAP + Anthropic = processo financeiro mais inteligente. 💡 Análise de variação, fechamento automatizado, projeção de caixa. Com dado real do seu SAP. #SAPAnthropic #FP&A #IA #SAP #Solveplan",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-2.md",
        "Fran", "🟡 Criado"
    ),
    (
        "12/06 (Sex)", "Sexta",
        "Case / Prova Social",
        "Clip — Case Lins Agroindustrial (vídeo)",
        "Case de cliente — vídeo",
        "Clip do depoimento: como a Lins Agroindustrial saiu de Excel para SAP Analytics Cloud",
        "Credibilidade / Geração de pipeline",
        "CFO, Controller, Head de Dados — Agro e Indústria",
        "Meio / Fundo",
        "[Ver legendas em: marketing/videos/clips/case-lins-agroindustrial/legendas.md] [Ver calendário de clips: marketing/videos/clips/case-lins-agroindustrial/calendario.md]",
        "[Ver legendas no arquivo de clips]",
        "[Ver legendas no arquivo de clips]",
        "Vídeo / Clip", "10h00",
        "marketing/videos/clips/case-lins-agroindustrial/",
        "Fran", "✅ Pronto"
    ),
    (
        "15/06 (Seg)", "Segunda",
        "Autoridade / Tendência",
        "Reforma tributária 2026 + arquitetura de dados SAP",
        "Criação própria — Gap de mercado único",
        "A reforma tributária de 2026 não é um problema de contabilidade. É um problema de arquitetura de dados.",
        "Geração de urgência / Awareness",
        "CFO, Controller, Diretor Fiscal, CIO",
        "Topo / Meio",
        "A reforma tributária de 2026 não é um problema de contabilidade. É um problema de arquitetura de dados. Novos regimes tributários, novas frequências de apuração, obrigações acessórias que exigem rastreabilidade completa — tudo isso pressiona a qualidade dos dados na fonte. Não dá para cumprir as obrigações de 2026 com a arquitetura de dados de 2019. O que as empresas que se antecipam estão fazendo: → Revisando dado mestre → Garantindo rastreabilidade em transações fiscais → Integrando fontes não-SAP com trilha de auditoria. Quem não estrutura vai cumprir manualmente, caro e tarde. Se você está nesse processo com SAP, fala comigo. #ReformaTributaria #SAP #DadosFiscais #Governança #SAPBusinessDataCloud #Solveplan",
        "Reforma tributária de 2026 não é problema de contabilidade — é problema de arquitetura de dados. Novas obrigações = nova exigência de rastreabilidade. Quem não estrutura o ambiente de dados cumpre manualmente, caro e tarde. #ReformaTributaria #SAP #DadosFiscais #Governança",
        "A reforma tributária 2026 exige mais do que mudanças contábeis. Exige dado rastreável, desde a fonte. 📋 Você está preparado? #ReformaTributaria #SAP #Dados #Governança #Solveplan",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-3.md",
        "Fran", "🟡 Criado"
    ),
    (
        "16/06 (Ter)", "Terça",
        "Artigo / Deep-dive",
        "SAP Knowledge Graph — IA que entende o negócio",
        "Artigo do blog",
        "Há uma diferença entre uma IA que sabe o que é reconciliação financeira e uma que sabe como ela funciona na sua empresa.",
        "Educação / Autoridade",
        "CIO, Diretor de TI, Head de Dados, CFO",
        "Topo / Meio",
        "Há uma diferença entre uma IA que sabe o que é reconciliação financeira e uma que sabe como ela funciona na sua empresa. Essa diferença tem nome: SAP Knowledge Graph. 452.000 tabelas do S/4HANA mapeadas com relações semânticas, processos e políticas. Para que os agentes naveguem no mapa do negócio — não no texto. E há uma implicação direta: o Knowledge Graph sabe como o ERP funciona em geral. O SAP BDC é o que faz ele saber como o seu ERP funciona especificamente. Escrevemos um artigo explicando o funcionamento e o que muda para quem avalia agentes SAP. → Link no primeiro comentário. #SAPKnowledgeGraph #IA #SAP #SAPBusinessDataCloud #FP&A #Solveplan",
        "IA que sabe o que é fechamento ≠ IA que sabe como o fechamento funciona na sua empresa. SAP Knowledge Graph: 452.000 tabelas mapeadas com semântica de processo. O BDC é o que faz esse mapa ser o mapa da sua empresa. Artigo completo → [link] #SAPKnowledgeGraph #IA #SAP",
        "IA sem contexto do negócio erra com precisão. 🎯 O SAP Knowledge Graph resolve: 452k tabelas do S/4 mapeadas semanticamente. Artigo: link na bio. #SAPKnowledgeGraph #IA #SAP #Dados",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-3.md | marketing/blog/sap-knowledge-graph/artigo.md",
        "Fran", "🟡 Criado"
    ),
    (
        "17/06 (Qua)", "Quarta",
        "Educação / Produto",
        "5 dores de dados que travam o FP&A",
        "Criação própria",
        "5 dores de dados que travam o FP&A — e o que o SAP BDC faz por cada uma",
        "Educação / Consideração de produto",
        "CFO, Controller, Head de Planejamento",
        "Meio",
        "5 dores de dados que travam o FP&A — e o que o SAP BDC faz por cada uma. 1) Ciclo de fechamento longo demais 2) Planejamento em Excel 3) Dado que existe mas ninguém confia 4) Sem visibilidade de custos de plataforma 5) IA sem fundação de dados pronta. Qual dessas está mais presente no seu dia a dia? Comenta abaixo. #FP&A #SAPBusinessDataCloud #Dados #Planejamento #Analytics #SAP #Solveplan",
        "5 dores que travam o FP&A — e o que o SAP BDC faz por cada uma. 1) Fechamento longo 2) Planejamento em Excel 3) Dado não confiável 4) Sem visibilidade de custos 5) IA sem dados prontos. Qual é a sua? → [link] #FP&A #SAPBusinessDataCloud #Dados",
        "5 dores que travam o FP&A. Desliza e me conta qual é a sua 👆 #FP&A #SAPBusinessDataCloud #Dados #Planejamento #SAP",
        "Carrossel (7 slides)", "10h00",
        "marketing/posts/junho-2026/semana-3.md",
        "Fran", "🟡 Criado"
    ),
    (
        "18/06 (Qui)", "Quinta",
        "Autoridade / Tendência",
        "Governança de dados é responsabilidade do CFO",
        "Criação própria",
        "Governança de dados não é um projeto de TI. É uma decisão estratégica do CFO.",
        "Autoridade / Posicionamento executivo",
        "CFO, Diretor Financeiro, Controller",
        "Topo / Meio",
        "Governança de dados não é um projeto de TI. É uma decisão estratégica do CFO. Ainda existe uma divisão equivocada: TI cuida dos dados, financeiro cuida dos números. O problema: os números saem dos dados. E quando os dados não têm governança, o financeiro passa o dia reconciliando inconsistências — não tomando decisões. O que governança significa para finanças: saber qual é a versão correta de um número, ter rastreabilidade, garantir que o dado de planejamento e o de fechamento vêm da mesma fonte. O SAP BDC é a plataforma onde esse argumento se torna implementação. Mas a decisão de priorizar governança precisa vir de quem entende o custo real da falta dela. #Governança #FP&A #CFO #Dados #SAP #SAPBusinessDataCloud #Solveplan",
        "Governança de dados não é projeto de TI — é decisão estratégica do CFO. Quando o dado não tem governança, o financeiro reconcilia em vez de decidir. O custo não aparece no P&L. Aparece no retrabalho e na auditoria. #Governança #FP&A #CFO #SAP",
        "Governança de dados não é TI. É decisão do CFO. 📊 Sem ela, o financeiro reconcilia em vez de decidir. #Governança #FP&A #CFO #SAP #Solveplan",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-3.md",
        "Fran", "🟡 Criado"
    ),
    (
        "19/06 (Sex)", "Sexta",
        "Autoridade / Tendência",
        "O que os melhores CFOs fazem diferente com dados em 2026",
        "Criação própria — Padrão observado em clientes",
        "Os CFOs que mais avançam em 2026 não têm as melhores planilhas. Têm os melhores dados.",
        "Autoridade / Geração de interesse",
        "CFO, Diretor Financeiro, Controller",
        "Topo",
        "Os CFOs que mais avançam em 2026 não têm as melhores planilhas. Têm os melhores dados. 5 comportamentos que observamos nos clientes que mais evoluíram: 1) Tratam dado mestre como ativo financeiro 2) Fecham o mês com dado — não com dado que vai ser ajustado depois 3) Simulam antes de comprometer 4) Sabem onde está o risco antes do conselho perguntar 5) Tratam a plataforma de dados como responsabilidade de finanças — não de TI. Esses comportamentos não surgem do nada. Surgem quando o ambiente técnico suporta — e quando o líder de finanças decide que dado é responsabilidade sua. #CFO #FP&A #Dados #Analytics #Planejamento #SAP #Solveplan",
        "5 comportamentos dos CFOs que mais avançam em 2026: 1) Dado mestre como ativo financeiro 2) Fechamento com dado validado 3) Cenários antes de comprometer 4) Risco visível antes do conselho 5) Dados como responsabilidade de finanças. O que todos têm em comum: um ambiente técnico que suporta. #CFO #FP&A #Dados #SAP",
        "CFOs que mais avançam em 2026: 5 comportamentos em comum. Desliza 👆 #CFO #FP&A #Dados #Analytics #SAP #Solveplan",
        "Texto longo", "8h30",
        "marketing/posts/junho-2026/semana-3.md",
        "Fran", "🟡 Criado"
    ),
    (
        "22/06 (Seg)", "Segunda",
        "Artigo / Deep-dive",
        "SAP Joule Work — passou a executar, não só sugerir",
        "Artigo do blog",
        "O SAP Joule deixou de ser uma caixa de perguntas. Passou a ser quem executa o processo.",
        "Educação / Autoridade",
        "CIO, Head de Dados, Gestor SAP, Analista Financeiro",
        "Meio",
        "O SAP Joule deixou de ser uma caixa de perguntas. Passou a ser quem executa o processo. → 'Prepare o relatório de variação de custo do trimestre para as plantas do Brasil' — o Joule busca, contextualiza e apresenta. Sem abrir um único menu SAP. → Lançamentos, aprovações, workflows — via linguagem natural, dentro das alçadas corretas, com trilha de auditoria. Para quem usa SAP Datasphere: o Joule está integrado. Um analista que antes navegava por espaços e views agora simplesmente pergunta. A SAP reconstruiu o Joule do zero após admitir que o Studio original não entregou o que prometia. → Link no primeiro comentário. #SAPJoule #SAPDatasphere #SAPBusinessDataCloud #Automação #IA #SAP #Solveplan",
        "O SAP Joule passou a executar processos — não só sugerir. Relatório de variação? Busca e apresenta. Aprovações? Executa dentro das alçadas. Tudo com linguagem natural e trilha de auditoria. Integrado ao SAP Datasphere. Artigo → [link] #SAPJoule #SAPDatasphere #IA #SAP",
        "O Joule agora executa — não só pergunta. 🤖 Relatórios, aprovações, análises via linguagem natural. Integrado ao SAP Datasphere. Artigo: link na bio. #SAPJoule #SAPDatasphere #IA #SAP",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-4.md | marketing/blog/sap-joule-work/artigo.md",
        "Fran", "🟡 Criado"
    ),
    (
        "23/06 (Ter)", "Terça",
        "Engagement",
        "Poll — Estágio SAP de dados",
        "Criação própria",
        "Sua empresa usa SAP. Em qual estágio está a estratégia de dados hoje?",
        "Engajamento / Geração de dados de audiência",
        "CFO, Controller, Head de Dados, CIO",
        "Topo",
        "Sua empresa usa SAP. Em qual estágio está a estratégia de dados hoje? A) Excel como ferramenta de análise — SAP é sistema de transação, planilha é o BI B) Usando Datasphere — saímos do Excel, ainda estruturando C) Implementando ou avaliando SAP BDC — caminhando para dado governado com IA D) Avaliando o próximo passo — temos SAP mas não sabemos por onde começar. O estágio mais comum que encontramos em diagnósticos é A ou D. Onde você está? Vote abaixo. #SAP #Dados #FP&A #Analytics #Planejamento #Solveplan",
        "Em qual estágio SAP de dados você está? A: Excel como BI | B: Usando Datasphere | C: Implementando BDC | D: Tem SAP mas não sabe por onde começar. O mais comum que vemos: A ou D. Onde você está? 👇 #SAP #Dados #FP&A #Analytics",
        "Sua empresa usa SAP. Qual o estágio de dados? 🔢 A: Excel | B: Datasphere | C: BDC | D: Avaliando. Conta nos comentários! #SAP #Dados #FP&A",
        "Poll (enquete)", "12h00",
        "marketing/posts/junho-2026/semana-4.md",
        "Fran", "🟡 Criado"
    ),
    (
        "24/06 (Qua)", "Quarta",
        "Educação / Produto",
        "Jornada SAP BDC — do assessment ao resultado",
        "Criação própria",
        "Da implementação SAP ao SAP BDC: o que muda em cada etapa — e o que esperar",
        "Consideração / Intenção de compra",
        "CFO, CIO, Diretor de Projetos SAP",
        "Meio / Fundo",
        "Da implementação SAP ao SAP BDC — o que acontece em cada etapa: 1) Assessment (4-6 semanas): diagnóstico do ambiente e roadmap priorizado 2) Fundação de dados (3-6 meses): Datasphere + dado mestre + governança 3) Ativação do BDC (2-4 meses): semântica, Insight Apps, integração com IA 4) Resultado sustentável: AMS + melhoria contínua. A jornada não precisa ser feita de uma vez. O que não pode acontecer é não começar. Em qual etapa você está hoje? #SAPBusinessDataCloud #Analytics #Dados #SAP #FP&A #Solveplan",
        "Jornada SAP BDC em 4 etapas: Assessment → Fundação de dados → Ativação BDC → Resultado sustentável. Tempos médios, entregáveis e o que esperar em cada fase. Em qual etapa você está? → [link] #SAPBusinessDataCloud #Analytics #SAP #Dados",
        "4 etapas para ir do SAP ao BDC. Desliza e veja em qual você está 👆 #SAPBusinessDataCloud #Analytics #SAP #FP&A #Solveplan",
        "Carrossel (6 slides)", "10h00",
        "marketing/posts/junho-2026/semana-4.md",
        "Fran", "🟡 Criado"
    ),
    (
        "25/06 (Qui)", "Quinta",
        "Case / Prova Social",
        "Case Vale — SAPPHIRE 2026 — 50% menos tempo de reporting",
        "Evento SAP SAPPHIRE 2026 — case público",
        "A Vale apresentou no SAPPHIRE 2026 um número que vale mais do que qualquer apresentação de produto.",
        "Credibilidade / Urgência",
        "CFO, Controller, Diretor Financeiro, CIO",
        "Meio / Fundo",
        "A Vale apresentou no SAPPHIRE 2026: 100+ planilhas Excel. 174 mil funcionários. 20 países. 24 mil centros de custo. Resultado depois da implementação SAP EPM: ~50% de redução no tempo de reporting financeiro. A Vale não tinha falta de dado. Tinha excesso de dado sem governança. 100+ planilhas = 100+ versões do mesmo número. A resposta: nenhuma. Migraram para um ambiente governado, com uma versão única — sempre atualizada, sempre rastreável. 50% menos tempo de reporting = semanas de volta para tomada de decisão. Se a maior mineradora da América Latina chegou lá — o que você está esperando? #SAP #FP&A #Reporting #Dados #Vale #SAPBusinessDataCloud #Solveplan",
        "A Vale no SAPPHIRE 2026: 100+ planilhas → -50% no tempo de reporting. 174k funcionários. 20 países. 24k centros de custo. O problema não era falta de dado — era excesso sem governança. Uma versão única resolveu. E você, o que está esperando? #SAP #FP&A #Vale #SAPPHIRE2026 #Dados",
        "Vale: 100+ planilhas → -50% no reporting. 📉 SAPPHIRE 2026. Governança de dados que gera resultado real. #SAP #FP&A #Vale #SAPPHIRE2026 #Dados #Solveplan",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-4.md",
        "Fran", "🟡 Criado"
    ),
    (
        "26/06 (Sex)", "Sexta",
        "Educação / Nicho",
        "Consolidação societária no SAP BDC",
        "Criação própria — Gap único de mercado",
        "Consolidação societária é um dos processos financeiros mais complexos de automatizar — e um dos que mais têm a ganhar com o SAP BDC.",
        "Geração de pipeline / Nicho",
        "CFO de grupo empresarial, Controller de consolidação, Diretor Financeiro corporativo",
        "Meio / Fundo",
        "Consolidação societária é um dos processos financeiros mais complexos de automatizar. E um dos que mais têm a ganhar com o SAP BDC. Entidades em países diferentes. Moedas diferentes. Planos de contas diferentes. Transações intercompany que precisam ser eliminadas. O que a maioria usa: planilhas. Com ajuste manual. O que o BDC muda: → Semântica de grupo: hierarquias mapeadas que o sistema entende → Eliminação de intercompany com rastreabilidade → Moeda funcional e de consolidação separadas sem retrabalho → Base integrada para SAP Group Reporting e SAP CFIN. Para CFOs de grupo: a diferença entre saber como está o grupo em dias ou em semanas. Se você cuida de consolidação societária em SAP, fala comigo. #Consolidação #FP&A #SAP #CFO #GrupoEmpresarial #SAPBusinessDataCloud #Solveplan",
        "Consolidação societária: o processo financeiro mais complexo de automatizar — e um dos que mais ganham com SAP BDC. Entidades, moedas, intercompany, regulatórios. A maioria ainda usa planilhas com ajuste manual. O BDC muda isso. CFOs de grupo: dias vs. semanas para fechar. #Consolidação #FP&A #SAP #CFO #GrupoEmpresarial",
        "Consolidação societária + SAP BDC = fechamento em dias, não semanas. 🏢 Para CFOs de grupo empresarial. #Consolidação #FP&A #SAP #CFO #SAPBusinessDataCloud #Solveplan",
        "Texto longo", "8h30",
        "marketing/posts/junho-2026/semana-4.md",
        "Fran", "🟡 Criado"
    ),
    (
        "29/06 (Seg)", "Segunda",
        "Artigo / Deep-dive",
        "SAP BDC como Knowledge Core — mapa genérico vs. mapa da empresa",
        "Artigo do blog",
        "Um agente SAP toma decisões com base no que sabe sobre o seu negócio — não sobre o SAP em geral.",
        "Educação / Autoridade",
        "CIO, Head de Dados, CFO, Arquiteto SAP",
        "Topo / Meio",
        "Um agente SAP toma decisões com base no que sabe sobre o seu negócio. Não sobre o SAP em geral. O SAP Knowledge Graph sabe como o ERP funciona — contexto genérico, compartilhado por todos. O SAP BDC é o que transforma esse mapa genérico no mapa da sua empresa. Sem o BDC: o agente sabe o que é um lançamento financeiro no S/4HANA. Com o BDC: ele sabe que na sua empresa esse lançamento precisa passar pelo controller regional se for acima de R$ 50k, que a empresa Brasil fecha no dia 5 e que o centro de custo 1200 pertence à unidade X. Sem esse contexto, a automação funciona. Mas não se pode confiar nela. → Link no primeiro comentário. #SAPBusinessDataCloud #KnowledgeGraph #IA #Agentes #SAP #FP&A #Solveplan",
        "SAP BDC como Knowledge Core: o mapa genérico vs. o mapa da sua empresa. Knowledge Graph sabe o SAP em geral. BDC faz ele saber o seu SAP especificamente. Sem isso: automação que funciona mas não se confia. Artigo → [link] #SAPBusinessDataCloud #KnowledgeGraph #IA #SAP",
        "Agente SAP precisa conhecer o seu negócio — não o SAP em geral. 🗺️ O BDC é o que faz essa diferença. Artigo: link na bio. #SAPBusinessDataCloud #KnowledgeGraph #IA #SAP",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-5.md | marketing/blog/sap-bdc-knowledge-core/artigo.md",
        "Fran", "🟡 Criado"
    ),
    (
        "30/06 (Ter)", "Terça",
        "Autoridade / Retrospectiva",
        "Encerramento de junho — O dado está pronto para decidir?",
        "Criação própria",
        "Em junho, o mercado SAP no Brasil teve uma conversa que precisava acontecer.",
        "Autoridade / Posicionamento de marca",
        "CFO, CIO, Head de Dados, Controller",
        "Topo",
        "Em junho, o mercado SAP no Brasil teve uma conversa que precisava acontecer. Não foi sobre migrar para a nuvem. Foi sobre uma pergunta mais fundamental: o seu dado está pronto para tomar decisão? 4 coisas que o mês nos mostrou: 1) Empresas pós-S/4HANA não sabem o que fazer com o dado do sistema 2) Reforma tributária pressiona governança de baixo para cima 3) IA sobre dado sem governança = automação de erros em escala 4) CFOs que assumem responsabilidade pelo dado avançam mais rápido. Em julho, seguimos com mais conteúdo. Se quer aprofundar algum tema ou entender como isso se aplica ao seu SAP, fala comigo. #SAP #Dados #FP&A #Analytics #SAPBusinessDataCloud #Solveplan",
        "Junho em resumo para o mercado SAP: o dado está pronto para tomar decisão? 4 aprendizados do mês: pós-S/4HANA sem direção clara, reforma tributária como urgência, IA sobre dado ruim = erro em escala, CFO que assume dado avança mais rápido. Em julho, seguimos. #SAP #Dados #FP&A #SAPBusinessDataCloud",
        "Junho em resumo: o dado está pronto para decidir? 🎯 4 coisas que o mês nos mostrou sobre o mercado SAP no Brasil. #SAP #Dados #FP&A #SAPBusinessDataCloud #Solveplan",
        "Texto longo", "9h00",
        "marketing/posts/junho-2026/semana-5.md",
        "Fran", "🟡 Criado"
    ),
]

# ── Escrever linhas ──────────────────────────────────────────────────────────
for row_idx, post in enumerate(posts, 2):
    is_even = (row_idx % 2 == 0)
    row_bg  = ROW_ALT if is_even else "FFFFFF"

    for col_idx, value in enumerate(post, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border    = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font      = Font(size=9)

        # cor de fundo base
        cell.fill = PatternFill("solid", fgColor=row_bg)

        # status colorido (última coluna = 17)
        if col_idx == 17:
            if "Pronto" in str(value):
                cell.fill = PatternFill("solid", fgColor=PRONTO_BG)
                cell.font = Font(size=9, bold=True, color=GREEN_FG)
            elif "Criado" in str(value):
                cell.fill = PatternFill("solid", fgColor=CRIADO_BG)
                cell.font = Font(size=9, bold=True, color=ORANGE_FG)

    ws.row_dimensions[row_idx].height = 80

# ── Larguras de coluna ───────────────────────────────────────────────────────
col_widths = {
    1: 14,   # Data
    2: 10,   # Dia
    3: 22,   # Linha Editorial
    4: 32,   # Tema
    5: 28,   # Origem
    6: 45,   # Título
    7: 28,   # Objetivo
    8: 32,   # Persona
    9: 16,   # Funil
    10: 70,  # Copy LinkedIn
    11: 55,  # Copy Facebook
    12: 45,  # Copy Instagram
    13: 18,  # Formato
    14: 10,  # Horário
    15: 55,  # Ref Texto
    16: 14,  # Responsável
    17: 14,  # Status
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Congelar cabeçalho e primeiras colunas ───────────────────────────────────
ws.freeze_panes = "C2"

# ── Filtro automático ────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── Salvar ───────────────────────────────────────────────────────────────────
output_path = r"c:\Users\franc\solveplan.com\Roberto Molina - Marketing\1. MKT Estrategy\3. Agentes de IA\ccos-ratos\dados\cronograma-posts-junho-2026.xlsx"
wb.save(output_path)
print(f"Arquivo salvo: {output_path}")

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

AZUL = "006AFF"
BRANCO = "FFFFFF"
CINZA_CLARO = "F5F5F5"
VERDE = "00C4A7"
LARANJA = "FF6B35"

def estilizar_aba(ws, df, cor_header=AZUL):
    header_fill = PatternFill("solid", fgColor=cor_header)
    header_font = Font(bold=True, color=BRANCO)
    alt_fill = PatternFill("solid", fgColor=CINZA_CLARO)
    borda = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    for i, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.border = borda
            cell.alignment = Alignment(vertical='center')
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical='center', horizontal='center')
            elif i % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


# ==================== TABELA 1: fLeads ====================
df_leads_raw = pd.read_excel('../dados/Leads_Geral.xlsx', sheet_name='Leads campanha')
fLeads = pd.DataFrame({
    'Data_Criacao': pd.to_datetime(df_leads_raw['Ceated_date'], errors='coerce'),
    'Ano': df_leads_raw['ano'].fillna(2026).astype('Int64'),
    'Quarter': df_leads_raw['Quarter'].fillna(''),
    'Campanha': df_leads_raw['Nome da Campanha'].fillna('Sem Campanha'),
    'Canal': df_leads_raw['Canal'].fillna('Outros'),
    'Origem': df_leads_raw['Origem'].fillna('Outros'),
    'Tipo_Canal': df_leads_raw['Canal'].apply(
        lambda x: 'Evento' if 'vent' in str(x).lower()
        else ('Digital' if str(x) in ['LinkedIN','Google'] else 'Site')
    ),
    'Qualificado': df_leads_raw['Qualificado'].fillna('Lead'),
    'Empresa': df_leads_raw['Company name'].fillna(''),
    'Cargo': df_leads_raw['Cargo'].fillna('Outros'),
    'Departamento': df_leads_raw['Departamento'].fillna('Outros'),
    'E_Cliente': df_leads_raw['E Cliente?'].fillna('Nao') if 'E Cliente?' in df_leads_raw.columns else df_leads_raw.get('É Cliente?', pd.Series(['Não']*len(df_leads_raw))).fillna('Não'),
})
fLeads['Mes_Num'] = fLeads['Data_Criacao'].dt.month
fLeads['Mes_Nome'] = fLeads['Data_Criacao'].dt.strftime('%b').fillna('')
fLeads['Data_Criacao'] = fLeads['Data_Criacao'].dt.date


# ==================== TABELA 2: fNegocios ====================
df_hub_raw = pd.read_excel('../dados/hubspot-crm-exports-todos-os-negocios-2026-06-08.xlsx')

etapas_map = {
    'contractsent': 'Negocio Ganho',
    'closedlost': 'Negocio Perdido',
    '1003801450': 'Negocio Perdido',
    '1329065912': 'Negocio Perdido',
    '1243950468': 'Discovery / Escopo',
    'appointmentscheduled': 'Proposta Enviada',
    'qualifiedtobuy': 'Negociacao',
    '1245098328': 'Proposta Tecnica e Escopo',
    '1245098329': 'Aprovacao Precificacao',
    '978106239': 'Conexao / Alinhamento Inicial',
    '1198208897': 'OnGoing',
    '951905607': 'Conta Mapeada',
    '1252470125': 'Conectado',
    '1361438855': 'Conectado',
    '1213830320': 'Enviado para Vendas',
    '1329066022': 'Enviado para Vendas',
    '952093258': 'Em Cadencia',
    '1329065907': 'Em Cadencia',
    '1198074992': 'Handover',
    '1198208898': 'Pre Renovacao',
    '1198208899': 'Renovacao',
    '1250627210': 'Negocio Perdido',
}

ordem_funil = [
    'Conexao / Alinhamento Inicial',
    'Enviado para Vendas',
    'Em Cadencia',
    'Conta Mapeada',
    'Conectado',
    'Discovery / Escopo',
    'Proposta Tecnica e Escopo',
    'Proposta Enviada',
    'Aprovacao Precificacao',
    'Negociacao',
    'OnGoing',
    'Pre Renovacao',
    'Renovacao',
    'Handover',
    'Negocio Ganho',
    'Negocio Perdido',
]

# Valores reais dos deals (via HubSpot MCP)
valores_reais = {
    'COPEL - Planejamento Integrado': 4446346.74,
    'Klabin - Extensao do Hypercare': 540360.00,
    'Klabin - Extensão do Hypercare': 540360.00,
    'Aegea - SIS GPI_ Planej Capex': 534638.00,
    'Aegea - Squad de Recursos - Marco/26 a Jun/26': 427728.00,
    'Aegea - Squad de Recursos - Março/26 à Jun/26': 427728.00,
    'ACHE - Licenciamento SAP Datapshere': 270000.00,
    'Real Vidros Distribuidora | Suporte e Fabrica': 150000.00,
    'Real Vidros Distribuidora | Suporte e Fábrica': 150000.00,
    'Klabin - Horas de Fabrica': 147000.00,
    'Klabin - Horas de Fábrica': 147000.00,
    'ALPARGATAS - Suporte SAC': 132000.00,
    'VALE - Suporte Consolidacao': 102400.00,
    'VALE - Suporte Consolidação': 102400.00,
    'IPIRANGA AGROINDUSTRIAL S/A': 33583.68,
    'CPFL - Planejamento de OPEX e CAPEX': 1500000.00,
    'SUPLEY - Planejamento Integrado': 1000000.00,
    'Axia Energia (Eletrobras) - Planejamento de Opex e RH': 600000.00,
    'Andrade Gutierrez - Planejamento de Opex e RH': 578000.00,
    'Grupo Aguas do Brasil - Planejamento de RG': 548000.00,
    'Grupo Águas do Brasil - Planejamento de RG': 548000.00,
    'Orbia (Amanco) - Planejamento de Opex e Capex': 400000.00,
    'SUPLEY - Fabrica de Analytics': 350000.00,
    'SUPLEY - Fábrica de Analytics': 350000.00,
    'Aegea - Squad de Recursos - Março/26 à Jun/26': 427728.00,
    'Klabin - Extensão do Hypercare': 540360.00,
    'Coopercitrus': 0,
}

fNegocios = pd.DataFrame({
    'Nome_Negocio': df_hub_raw['Nome do negocio'] if 'Nome do negocio' in df_hub_raw.columns else df_hub_raw.iloc[:, 1],
    'Data_Criacao': pd.to_datetime(df_hub_raw['Data de criacao'] if 'Data de criacao' in df_hub_raw.columns else df_hub_raw.iloc[:, 2], errors='coerce'),
    'Data_Fechamento': pd.to_datetime(df_hub_raw['Data de fechamento'] if 'Data de fechamento' in df_hub_raw.columns else df_hub_raw.iloc[:, 4], errors='coerce'),
    'Proprietario': df_hub_raw.iloc[:, 3],
    'Campanha_Conversao': df_hub_raw.iloc[:, 5].fillna('Sem Campanha'),
    'Etapa_Original': df_hub_raw.iloc[:, 6],
    'Origem': df_hub_raw.iloc[:, 7].fillna('Outros'),
    'Empresa': df_hub_raw.iloc[:, 8].fillna(''),
})

# nome correto da coluna
col_nome = df_hub_raw.columns[1]
col_etapa = df_hub_raw.columns[6]
col_origem = df_hub_raw.columns[7]
col_campanha = df_hub_raw.columns[5]
col_empresa = df_hub_raw.columns[8]
col_data = df_hub_raw.columns[2]
col_fecha = df_hub_raw.columns[4]
col_prop = df_hub_raw.columns[3]

fNegocios = pd.DataFrame({
    'Nome_Negocio': df_hub_raw[col_nome],
    'Data_Criacao': pd.to_datetime(df_hub_raw[col_data], errors='coerce'),
    'Data_Fechamento': pd.to_datetime(df_hub_raw[col_fecha], errors='coerce'),
    'Proprietario': df_hub_raw[col_prop],
    'Campanha_Conversao': df_hub_raw[col_campanha].fillna('Sem Campanha'),
    'Etapa_ID': df_hub_raw[col_etapa].astype(str),
    'Origem': df_hub_raw[col_origem].fillna('Outros'),
    'Empresa': df_hub_raw[col_empresa].fillna(''),
})

fNegocios['Etapa'] = fNegocios['Etapa_ID'].map(etapas_map).fillna(fNegocios['Etapa_ID'])
fNegocios['Status'] = fNegocios['Etapa'].apply(
    lambda x: 'Ganho' if x == 'Negocio Ganho'
    else ('Perdido' if x == 'Negocio Perdido' else 'Ativo')
)
fNegocios['Ano'] = fNegocios['Data_Criacao'].dt.year
fNegocios['Quarter'] = fNegocios['Data_Criacao'].dt.quarter.apply(lambda x: f'Q{int(x)}' if pd.notna(x) else '')
fNegocios['Mes_Num'] = fNegocios['Data_Criacao'].dt.month
fNegocios['Mes_Nome'] = fNegocios['Data_Criacao'].dt.strftime('%b').fillna('')
fNegocios['Valor_BRL'] = fNegocios['Nome_Negocio'].map(valores_reais).fillna(0)
fNegocios['Ordem_Funil'] = fNegocios['Etapa'].map({e: i for i, e in enumerate(ordem_funil)}).fillna(99).astype(int)
fNegocios['Data_Criacao'] = fNegocios['Data_Criacao'].dt.date
fNegocios['Data_Fechamento'] = fNegocios['Data_Fechamento'].dt.date
fNegocios = fNegocios.drop(columns=['Etapa_ID'])


# ==================== TABELA 3: fSessoes_GA4 ====================
fSessoes = pd.DataFrame({
    'Canal_Origem': [
        'linkedin / paidsocial', 'linkedin / social', '(direct) / (none)',
        'google / organic', 'LinkedInAds / paid', 'linkedin / paid',
        'bing / organic', 'email / e-mail', 'google / cpc',
        'facebook.com / referral', 'Outros canais'
    ],
    'Canal_Agrupado': [
        'LinkedIn Ads', 'LinkedIn Organico', 'Direto',
        'Google Organico', 'LinkedIn Ads', 'LinkedIn Ads',
        'Bing Organico', 'Email', 'Google Ads',
        'Facebook', 'Outros'
    ],
    'Tipo': [
        'Pago', 'Organico', 'Direto',
        'Organico', 'Pago', 'Pago',
        'Organico', 'Email', 'Pago',
        'Social', 'Outros'
    ],
    'Sessoes': [8296, 3485, 1737, 1441, 664, 523, 150, 71, 28, 21, 724],
    'Periodo': ['Jan-Mai 2026'] * 11,
    'Ano': [2026] * 11,
})


# ==================== TABELA 4: fGoogleAds ====================
fGoogleAds = pd.DataFrame({
    'Campanha': ['[Leads-Search 2026] Download e-book SAC Planning'] * 6,
    'Keyword': [
        'dados e IA', 'recursos de BI', 'orcamentos',
        'planejamento financeiro', 'excel planilhas', 'planilhas excel'
    ],
    'Cliques': [210, 144, 70, 53, 17, 16],
    'Impressoes': [911, 1768, 359, 141, 1178, 334],
    'CPC_BRL': [2.93, 2.34, 1.42, 1.18, 3.91, 4.08],
    'Custo_BRL': [614.51, 336.73, 99.46, 62.55, 66.40, 65.24],
    'Conversoes': [12, 6, 2, 2, 2, 1],
    'CTR_Perc': [23.05, 8.15, 19.50, 37.59, 1.44, 4.79],
    'Periodo': ['Jan-Mai 2026'] * 6,
})
fGoogleAds['Custo_por_Conversao'] = (fGoogleAds['Custo_BRL'] / fGoogleAds['Conversoes']).round(2)


# ==================== TABELA 5: fRedesSociais ====================
fRedesSociais = pd.DataFrame({
    'Plataforma': [
        'Instagram','Instagram','Instagram','Instagram','Instagram',
        'Instagram','Instagram','Instagram','Instagram','Instagram',
        'Facebook','Facebook','Facebook','Facebook','Facebook',
    ],
    'Tipo_Post': [
        'Post','Post','Post','Post','Post','Post','Post','Post','Reel','Reel',
        'Post','Post','Post','Post','Post',
    ],
    'Descricao': [
        'Melhor Parceiro SAP Award 2026','Planejamento financeiro em silos',
        'SAP Sapphire 2026','Patrocinadora BTP Experience','SAP BTP Experience anuncio',
        'SolvePlan na midia','SAP Sapphire artigo','SAP Partner Award reconhecimento',
        'Planejamento financeiro Reel','SAP Sapphire Reel',
        'Planejamento financeiro FB','SolvePlan patrocinadora v1',
        'SolvePlan patrocinadora v2','SolvePlan na midia FB','Melhor Parceiro SAP BDC FB',
    ],
    'Alcance': [300,202,156,109,104,90,87,71,None,None,6,5,9,4,5],
    'Curtidas_Likes': [40,18,2,8,4,3,6,14,18,2,None,None,None,None,None],
    'Comentarios': [3,0,1,0,0,1,0,1,0,1,None,None,None,None,None],
    'Taxa_Engajamento_Perc': [15.0,8.91,1.92,7.34,4.81,4.44,8.05,23.94,None,None,None,None,None,None,None],
    'Visualizacoes_Reel': [None,None,None,None,None,None,None,None,294,197,None,None,None,None,None],
    'Periodo': ['Jan-Mai 2026'] * 15,
})


# ==================== TABELA 6: dCampanha ====================
dCampanha = pd.DataFrame({
    'Campanha': [
        'SAP BTP Experience 2026', 'IT Summit Agro 2026', 'Formulario do Site',
        'Live Chat', 'Download e-book SAC', 'Expansao Cross/UpSell',
        'SAP NOW 2025', 'Reativacao', 'Indicacao SAP', 'Relacionamento AE',
        'Sem Campanha',
    ],
    'Tipo_Campanha': [
        'Evento SAP', 'Evento Proprio', 'Inbound', 'Inbound', 'Inbound',
        'Expansao Carteira', 'Evento SAP', 'Reativacao', 'Parceria', 'Carteira AE',
        'Sem Campanha',
    ],
    'Canal_Principal': [
        'Evento', 'Evento', 'Site', 'Site', 'Google Ads',
        'CRM', 'Evento', 'CRM', 'Indicacao', 'CRM',
        'N/A',
    ],
    'Investimento_BRL': [73385.91, 0, 0, 0, 1727.96, 0, 0, 0, 0, 0, 0],
    'Total_Leads': [176, 121, 5, 3, 2, 0, 0, 0, 0, 0, 0],
    'CPL_BRL': [416.97, 0, 0, 0, 863.98, 0, 0, 0, 0, 0, 0],
    'Periodo': [
        'Fev 2026', 'Abr 2026', 'Jan-Mai 2026', 'Jan-Mai 2026', 'Jan-Mai 2026',
        'Jan-Mai 2026', '2025', 'Jan-Mai 2026', 'Jan-Mai 2026', 'Jan-Mai 2026',
        'N/A',
    ],
})


# ==================== TABELA 7: KPIs_Resumo ====================
kpis = pd.DataFrame({
    'Metrica': [
        'Total Leads Gerados','MQLs','OPPs','Negócios Ganhos',
        'Receita Gerada (R$)','Pipeline Ativo (R$)','Ticket Médio (R$)',
        'Sessões Site','Usuários Ativos','Taxa de Rejeição (%)',
        'Investimento Google Ads (R$)','CTR Google Ads (%)','CPC Google Ads (R$)',
        'Custo BTP Experience (R$)','CPL BTP Experience (R$)','ROI BTP Experience (x)',
        'Seguidores Instagram','Taxa Engajamento Instagram (%)',
    ],
    'Valor': [
        307, 212, 9, 10,
        6784056.42, 22989571.50, 678405.64,
        17140, 13294, 78.84,
        1727.96, 4.97, 2.48,
        73385.91, 416.97, 20.0,
        423, 7.14,
    ],
    'Variacao_vs_Anterior': [
        None, None, None, None,
        None, None, None,
        '+160%', '+136%', '+32% (piora)',
        '-41%', '-39%', '+110%',
        None, None, None,
        '+4%', '+100%',
    ],
    'Status': [
        'OK','OK','OK','OK',
        'OK','OK','OK',
        'OK','OK','ATENCAO',
        'OK','OK','OK',
        'OK','OK','OK',
        'OK','OK',
    ],
    'Periodo': ['Jan-Mai 2026'] * 18,
})


# ==================== GERAR EXCEL ====================
output_path = '../dados/PowerBI_Solveplan_H1_2026.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    for df, nome in [
        (fLeads, 'fLeads'),
        (fNegocios, 'fNegocios'),
        (fSessoes, 'fSessoes_GA4'),
        (fGoogleAds, 'fGoogleAds'),
        (fRedesSociais, 'fRedesSociais'),
        (dCampanha, 'dCampanha'),
        (kpis, 'KPIs_Resumo'),
    ]:
        df.to_excel(writer, sheet_name=nome, index=False)
        ws = writer.sheets[nome]
        estilizar_aba(ws, df)

print(f'Arquivo gerado: {output_path}')
print('Abas criadas: fLeads, fNegocios, fSessoes_GA4, fGoogleAds, fRedesSociais, dCampanha, KPIs_Resumo')

"""
apollo_enriquecimento.py — Enriquece lista de empresas com contatos do Apollo.io

RODA NA MÁQUINA COM A API KEY DO APOLLO.

Uso:
    python scripts/apollo_enriquecimento.py dados/Empresas_Kevin_Enriquecido_v2_2026-05-24.xlsx

Dependências:
    pip install requests openpyxl

Configuração:
    Definir a variável de ambiente APOLLO_API_KEY antes de rodar:
    Windows:  set APOLLO_API_KEY=sua_chave_aqui
    Linux/Mac: export APOLLO_API_KEY=sua_chave_aqui
"""

import sys
import os
import time
import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import date

# ── Configuração ──────────────────────────────────────────────────────────────

APOLLO_API_KEY = os.environ.get("APOLLO_API_KEY", "")
APOLLO_BASE = "https://api.apollo.io/v1"

# Hierarquia de títulos P1–P8 (ver references/grupos-contato.md)
TITULOS_PRIORITARIOS = {
    "P1": ["CFO", "Chief Financial Officer"],
    "P2": ["Diretor Financeiro", "Diretora Financeira", "Finance Director",
            "Director of Finance", "Head of Finance", "VP Finance",
            "Vice President Finance", "Diretor Administrativo Financeiro"],
    "P3": ["CIO", "Chief Information Officer", "CTO", "Chief Technology Officer",
            "Diretor de TI", "Diretora de TI", "Diretor Tecnologia",
            "Diretora Tecnologia", "Diretor de Tecnologia", "IT Director",
            "Director of IT", "Technology Director", "Head of IT",
            "Head de TI", "Head of Technology", "Head Tecnologia"],
    "P4": ["Diretor de Sistemas", "Diretora de Sistemas", "Systems Director",
            "Diretor ERP", "Diretora ERP", "ERP Director", "SAP Director",
            "Diretor SAP", "Diretora SAP", "Head of Systems",
            "Head de Sistemas", "Head ERP", "Head SAP"],
    "P5": ["Gerente Financeiro", "Gerente Financeira", "Finance Manager",
            "Financial Manager", "Controller", "Controllership",
            "Gerente de Controladoria", "Controladoria", "FP&A",
            "Head of FP&A", "FP&A Manager", "Accounting Manager", "Gerente Contábil"],
    "P6": ["Gerente de TI", "Gerente TI", "IT Manager", "Technology Manager",
            "Gerente de Tecnologia", "Gerente Tecnologia",
            "Information Technology Manager"],
    "P7": ["Gerente de Sistemas", "Systems Manager", "Gerente Sistemas",
            "ERP Manager", "Gerente ERP", "SAP Manager", "Gerente SAP",
            "Gerente de Aplicações", "Applications Manager",
            "Gerente de Sistemas Corporativos"],
    "P8": ["Coordenador de TI", "Coordenadora de TI", "IT Coordinator",
            "Coordenador de Sistemas", "Coordenadora de Sistemas",
            "Coordenador SAP", "Coordenadora SAP", "Coordenador ERP"],
}

AREA_POR_PRIORIDADE = {
    "P1": "Financeiro", "P2": "Financeiro", "P5": "Financeiro",
    "P3": "TI", "P6": "TI",
    "P4": "Sistemas/ERP", "P7": "Sistemas/ERP", "P8": "Sistemas/ERP",
}

TERMOS_RUINS = [
    "intern", "estagi", "trainee", "aprendiz", "assistant", "assistente",
    "junior", " jr ", "analyst", "analista", "developer", "desenvolvedor",
    "sales", "vendas", "account executive", "business development",
    "new business", "marketing", "recruiter", "talent", "people",
    " hr ", "human resources", "customer success", "inside sales",
    " sdr", " bdr", "consultant", "consultor", "estágio",
    "jovem aprendiz", "auxiliar",
]

META_CONTATOS_POR_EMPRESA = 4  # 2 Financeiro + 2 TI/Sistemas


# ── Funções de classificação ──────────────────────────────────────────────────

def classificar_contato(titulo):
    """Retorna (prioridade, area) ou (None, None) se cargo ruim."""
    titulo_lower = titulo.lower()

    # Descartar termos ruins
    for termo in TERMOS_RUINS:
        if termo.lower() in titulo_lower:
            return None, None

    # Classificar por prioridade
    for prioridade, titulos in TITULOS_PRIORITARIOS.items():
        for t in titulos:
            if t.lower() in titulo_lower:
                area = AREA_POR_PRIORIDADE.get(prioridade, "Outro")
                return prioridade, area

    return None, None


def tem_cobertura_minima(contatos):
    """Verifica se já tem ao menos 1 Financeiro + 1 TI/Sistemas."""
    areas = [c.get("AREA") for c in contatos]
    return "Financeiro" in areas and ("TI" in areas or "Sistemas/ERP" in areas)


# ── Apollo API ────────────────────────────────────────────────────────────────

def extrair_cnpj_apollo(org):
    """Extrai CNPJ do objeto de organização retornado pelo Apollo."""
    # Apollo retorna CNPJ em raw_address, short_description ou campos customizados
    for campo in ["raw_address", "short_description", "seo_description"]:
        texto = org.get(campo, "") or ""
        matches = re.findall(r'\d{2}[\.\-]?\d{3}[\.\-]?\d{3}[\/\-]?\d{4}[\-]?\d{2}', texto)
        for m in matches:
            digits = re.sub(r'\D', '', m)
            if len(digits) == 14:
                return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"

    # Tentar via custom_fields
    for field in org.get("custom_fields", []):
        val = str(field.get("value", ""))
        matches = re.findall(r'\d{2}[\.\-]?\d{3}[\.\-]?\d{3}[\/\-]?\d{4}[\-]?\d{2}', val)
        for m in matches:
            digits = re.sub(r'\D', '', m)
            if len(digits) == 14:
                return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"

    return "pendente"


def buscar_empresa_apollo(nome, website=None):
    """Busca empresa no Apollo — retorna (id, nome, cnpj)."""
    endpoint = f"{APOLLO_BASE}/organizations/search"
    payload = {"q_organization_name": nome, "per_page": 1}
    if website and website not in ("N/D", "", None):
        domain = website.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0]
        payload["q_organization_domains"] = [domain]

    resp = _apollo_post(endpoint, payload)
    orgs = resp.get("organizations", [])
    if orgs:
        org = orgs[0]
        cnpj = extrair_cnpj_apollo(org)
        return org.get("id"), org.get("name"), cnpj
    return None, None, "pendente"


def buscar_contatos_apollo(org_id, org_name):
    """Busca contatos de uma empresa no Apollo filtrados por título."""
    endpoint = f"{APOLLO_BASE}/mixed_people/search"

    todos_contatos = []

    # Buscar por cada grupo de prioridade (P1 primeiro)
    for prioridade, titulos in TITULOS_PRIORITARIOS.items():
        if tem_cobertura_minima(todos_contatos) and len(todos_contatos) >= META_CONTATOS_POR_EMPRESA:
            break

        payload = {
            "q_organization_ids": [org_id],
            "person_titles": titulos[:5],  # Apollo aceita até 5 títulos por vez
            "per_page": 5,
            "page": 1,
        }

        resp = _apollo_post(endpoint, payload)
        pessoas = resp.get("people", [])

        for p in pessoas:
            titulo = p.get("title", "")
            p_class, area = classificar_contato(titulo)

            if not p_class:
                continue

            email = ""
            status = "partial"
            if p.get("email"):
                email = p["email"]
                status = "verified" if not p.get("email_status") == "invalid" else "low_confidence"

            linkedin = p.get("linkedin_url", "")
            telefone = ""
            if p.get("phone_numbers"):
                telefone = p["phone_numbers"][0].get("sanitized_number", "")

            contato = {
                "EMPRESA": org_name,
                "NOME_CONTATO": p.get("name", ""),
                "CARGO": titulo,
                "AREA": area,
                "PRIORIDADE_CONTATO": p_class,
                "EMAIL": email,
                "LINKEDIN_URL": linkedin,
                "TELEFONE": telefone,
                "FONTE": "Apollo",
                "STATUS_CONTATO": status,
                "OUTREACH": "",
            }
            todos_contatos.append(contato)

        time.sleep(0.5)

    # Ordenar por prioridade
    todos_contatos.sort(key=lambda x: x["PRIORIDADE_CONTATO"])
    return todos_contatos


def _apollo_post(endpoint, payload):
    """Faz POST na API do Apollo com retry simples."""
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "X-Api-Key": APOLLO_API_KEY,
    }
    for tentativa in range(3):
        try:
            resp = requests.post(endpoint, json=payload, headers=headers, timeout=15)
            if resp.status_code == 429:
                print("  Rate limit atingido, aguardando 30s...")
                time.sleep(30)
                continue
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            print(f"  Erro (tentativa {tentativa+1}): {e}")
            time.sleep(5)
    return {}


# ── Main ──────────────────────────────────────────────────────────────────────

def main(caminho_excel):
    if not APOLLO_API_KEY:
        print("ERRO: variável APOLLO_API_KEY não definida.")
        print("Execute: set APOLLO_API_KEY=sua_chave_aqui")
        sys.exit(1)

    path = Path(caminho_excel)
    if not path.exists():
        print(f"Arquivo não encontrado: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    idx_empresa = headers.index('EMPRESA') if 'EMPRESA' in headers else 0
    idx_cnpj = headers.index('CNPJ') if 'CNPJ' in headers else None
    idx_website = headers.index('WEBSITE') if 'WEBSITE' in headers else None

    empresas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        empresas.append({
            "nome": row[idx_empresa],
            "cnpj": row[idx_cnpj] if idx_cnpj else "",
            "website": row[idx_website] if idx_website else "",
        })

    total = len(empresas)
    print(f"\nProcessando {total} empresas no Apollo...\n")

    todos_contatos = []
    sem_contato = []

    for i, emp in enumerate(empresas, 1):
        nome = emp["nome"]
        print(f"[{i}/{total}] {nome}...", end=' ', flush=True)

        org_id, org_name, cnpj_apollo = buscar_empresa_apollo(nome, emp["website"])
        if not org_id:
            print("empresa não encontrada no Apollo")
            sem_contato.append(nome)
            continue

        # Preferir CNPJ do Apollo; fallback para o que já estava no Excel
        cnpj_final = cnpj_apollo if cnpj_apollo != "pendente" else emp["cnpj"]

        contatos = buscar_contatos_apollo(org_id, org_name or nome)
        for c in contatos:
            c["CNPJ"] = cnpj_final

        # Atualizar CNPJ no Excel de entrada também
        emp["cnpj"] = cnpj_final

        print(f"{len(contatos)} contatos encontrados")
        todos_contatos.extend(contatos)

        if not contatos:
            sem_contato.append(nome)

        time.sleep(1)

    # ── Gerar Excel de saída ──
    saida_path = path.parent / f"{path.stem}_contatos_{date.today().strftime('%Y-%m-%d')}.xlsx"
    wb_out = openpyxl.Workbook()

    # Aba 1 — empresas com CNPJ atualizado pelo Apollo
    ws_emp = wb_out.active
    ws_emp.title = "Empresas"
    emp_headers = [cell.value for cell in ws[1]]
    ws_emp.append(emp_headers)
    idx_cnpj_col = emp_headers.index('CNPJ') if 'CNPJ' in emp_headers else None
    emp_map = {e["nome"]: e["cnpj"] for e in empresas}
    idx_emp_col = emp_headers.index('EMPRESA') if 'EMPRESA' in emp_headers else 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        if idx_cnpj_col is not None:
            nome_row = row_list[idx_emp_col]
            row_list[idx_cnpj_col] = emp_map.get(nome_row, row_list[idx_cnpj_col])
        ws_emp.append(row_list)

    # Aba 2 — Contatos
    ws_con = wb_out.create_sheet("Contatos")
    col_headers = ["EMPRESA", "CNPJ", "NOME_CONTATO", "CARGO", "AREA",
                   "PRIORIDADE_CONTATO", "EMAIL", "LINKEDIN_URL", "TELEFONE",
                   "FONTE", "STATUS_CONTATO", "OUTREACH"]
    ws_con.append(col_headers)

    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws_con[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for c in todos_contatos:
        ws_con.append([c.get(h, "") for h in col_headers])

    # Larguras
    widths = [35, 22, 30, 35, 15, 18, 35, 45, 18, 10, 16, 60]
    for i, w in enumerate(widths, 1):
        ws_con.column_dimensions[ws_con.cell(1, i).column_letter].width = w

    wb_out.save(saida_path)

    # ── Resumo ──
    print(f"\n{'='*50}")
    print(f"RESUMO")
    print(f"{'='*50}")
    print(f"Empresas processadas: {total}")
    print(f"Total de contatos:    {len(todos_contatos)}")
    print(f"Sem contato:          {len(sem_contato)}")
    if sem_contato:
        print("\nEmpresas sem contato encontrado:")
        for e in sem_contato:
            print(f"  - {e}")

    # Breakdown por área
    areas = {}
    for c in todos_contatos:
        a = c.get("AREA", "Outro")
        areas[a] = areas.get(a, 0) + 1
    print("\nBreakdown por área:")
    for a, n in sorted(areas.items()):
        print(f"  {a}: {n}")

    verified = sum(1 for c in todos_contatos if c.get("STATUS_CONTATO") == "verified")
    print(f"\nContatos com email verified: {verified}/{len(todos_contatos)}")
    print(f"\nArquivo salvo: {saida_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/apollo_enriquecimento.py <caminho_excel>")
        sys.exit(1)
    main(sys.argv[1])

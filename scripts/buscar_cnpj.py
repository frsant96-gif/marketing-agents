"""
buscar_cnpj.py — Busca CNPJ por nome de empresa via Google (gratuito)

Uso:
    python scripts/buscar_cnpj.py dados/Empresas_Kevin_Enriquecido_v2_2026-05-24.xlsx

Dependências:
    pip install requests beautifulsoup4 openpyxl
"""

import sys
import time
import re
import requests
from bs4 import BeautifulSoup
import openpyxl
from pathlib import Path

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "pt-BR,pt;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

CNPJ_RE = re.compile(r'\d{2}[\.\-]?\d{3}[\.\-]?\d{3}[\/\-]?\d{4}[\-]?\d{2}')


def formatar_cnpj(raw):
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return raw


def validar_cnpj(cnpj):
    """Validação básica de dígitos verificadores."""
    digits = re.sub(r'\D', '', cnpj)
    if len(digits) != 14 or len(set(digits)) == 1:
        return False
    # Primeiro dígito verificador
    pesos = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    soma = sum(int(digits[i]) * pesos[i] for i in range(12))
    resto = soma % 11
    d1 = 0 if resto < 2 else 11 - resto
    if int(digits[12]) != d1:
        return False
    # Segundo dígito verificador
    pesos = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    soma = sum(int(digits[i]) * pesos[i] for i in range(13))
    resto = soma % 11
    d2 = 0 if resto < 2 else 11 - resto
    return int(digits[13]) == d2


def buscar_via_google(nome, razao=None):
    """Busca CNPJ no Google e extrai do snippet de resultado."""
    termos = [razao, nome]
    for termo in termos:
        if not termo:
            continue
        query = f'CNPJ "{termo}"'
        url = f"https://www.google.com/search?q={requests.utils.quote(query)}&num=5&hl=pt-BR"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=12)
            texto = resp.text

            # Extrair todos os CNPJs do HTML
            candidatos = CNPJ_RE.findall(texto)
            validos = []
            for c in candidatos:
                if validar_cnpj(c):
                    fmt = formatar_cnpj(c)
                    if fmt not in validos:
                        validos.append(fmt)

            if len(validos) == 1:
                return validos[0], "encontrado"
            elif len(validos) > 1:
                return validos[0], "ambíguo"

        except Exception as e:
            print(f"\n  Erro Google '{termo}': {e}")

        time.sleep(3)

    return "pendente", "não encontrado"


def buscar_via_receitaws(cnpj_raw):
    """Valida e enriquece CNPJ já encontrado via ReceitaWS."""
    digits = re.sub(r'\D', '', cnpj_raw)
    if len(digits) != 14:
        return None
    try:
        r = requests.get(f"https://receitaws.com.br/v1/cnpj/{digits}",
                         headers=HEADERS, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if data.get("status") != "ERROR":
                return formatar_cnpj(digits)
    except Exception:
        pass
    return formatar_cnpj(digits)


def main(caminho_excel):
    path = Path(caminho_excel)
    if not path.exists():
        print(f"Arquivo não encontrado: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    if 'CNPJ' not in headers:
        print("Coluna CNPJ não encontrada.")
        sys.exit(1)

    idx_empresa = headers.index('EMPRESA') if 'EMPRESA' in headers else None
    idx_razao = headers.index('RAZAO SOCIAL') if 'RAZAO SOCIAL' in headers else None
    idx_cnpj = headers.index('CNPJ')

    total = ws.max_row - 1
    print(f"\nProcessando {total} empresas via Google...\n")

    encontrados = ambiguos = pendentes = 0

    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        empresa = row[idx_empresa].value if idx_empresa is not None else ""
        razao = row[idx_razao].value if idx_razao is not None else ""
        atual = row[idx_cnpj].value

        if atual and atual not in ('pendente', 'ambíguo', '', None):
            print(f"[{i}/{total}] {empresa} — já tem CNPJ ({atual})")
            encontrados += 1
            continue

        print(f"[{i}/{total}] {empresa}...", end=' ', flush=True)
        cnpj, status = buscar_via_google(empresa, razao)
        row[idx_cnpj].value = cnpj
        print(f"{cnpj}")

        if status == "encontrado":
            encontrados += 1
        elif status == "ambíguo":
            ambiguos += 1
        else:
            pendentes += 1

        # Salvar progresso a cada 5 empresas
        if i % 5 == 0:
            wb.save(path)

        time.sleep(4)  # evitar bloqueio do Google

    wb.save(path)

    print(f"\n--- Resultado ---")
    print(f"Encontrados:  {encontrados}/{total}")
    print(f"Ambíguos:     {ambiguos}/{total} — verificar manualmente")
    print(f"Pendentes:    {pendentes}/{total} — verificar manualmente")
    print(f"Arquivo salvo: {path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/buscar_cnpj.py <caminho_excel>")
        sys.exit(1)
    main(sys.argv[1])

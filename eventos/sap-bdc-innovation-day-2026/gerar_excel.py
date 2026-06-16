import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AZUL = "002F6C"
AMARELO = "F5A800"
BRANCO = "FFFFFF"
CINZA_CLARO = "F4F4F4"
CINZA_MEDIO = "DDDDDD"
AZUL_CLARO = "E8F0FE"
VERDE = "1A7A4A"
VERMELHO = "C0392B"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=None, size=11, italic=False):
    return Font(bold=bold, color=color or "000000", size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, bg=None, fg="000000", h_align="left", size=11, italic=False, wrap=True, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = align(h_align, wrap=wrap)
    if bg:
        cell.fill = fill(bg)
    if border:
        cell.border = border_thin()
    return cell

wb = openpyxl.Workbook()

# ==========================================
# ABA 1 — CHECKLIST GERAL
# ==========================================
ws1 = wb.active
ws1.title = "Checklist Geral"

ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 40
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 16
ws1.column_dimensions["F"].width = 14

# Cabeçalho principal
ws1.merge_cells("A1:F1")
set_cell(ws1, 1, 1, "SAP BDC Innovation Day 2026 — Checklist Geral", bold=True, bg=AZUL, fg=BRANCO, h_align="center", size=14, border=False)
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:F2")
set_cell(ws1, 2, 1, "26 de agosto de 2026  |  SAP Brasil, São Paulo  |  08:30 às 12h", bg=AMARELO, fg=AZUL, h_align="center", size=11, bold=True, border=False)
ws1.row_dimensions[2].height = 20

# Header colunas
row = 4
headers = ["✓", "Atividade", "Responsável", "Prazo", "Status", "Observações"]
for i, h in enumerate(headers, 1):
    set_cell(ws1, row, i, h, bold=True, bg=AZUL, fg=BRANCO, h_align="center", size=11)
ws1.row_dimensions[row].height = 20
row += 1

def section_header(ws, row, title):
    ws.merge_cells(f"A{row}:F{row}")
    set_cell(ws, row, 1, title, bold=True, bg=AMARELO, fg=AZUL, h_align="left", size=11)
    ws.row_dimensions[row].height = 18
    return row + 1

def item(ws, row, atividade, responsavel="", prazo="", status="A fazer", obs="", zebra=False):
    bg = CINZA_CLARO if zebra else BRANCO
    set_cell(ws, row, 1, "☐", bg=bg, h_align="center")
    set_cell(ws, row, 2, atividade, bg=bg)
    set_cell(ws, row, 3, responsavel, bg=bg, h_align="center")
    set_cell(ws, row, 4, prazo, bg=bg, h_align="center")
    set_cell(ws, row, 5, status, bg=bg, h_align="center")
    set_cell(ws, row, 6, obs, bg=bg)
    ws.row_dimensions[row].height = 18
    return row + 1

# ---- PLANEJAMENTO E DEFINIÇÕES ----
row = section_header(ws1, row, "📋  PLANEJAMENTO E DEFINIÇÕES")
items_plan = [
    ("Confirmar reserva do Auditório MASP — SAP Brasil", "Fran", "20/06/2026"),
    ("Definir nome definitivo do evento", "Fran", "20/06/2026"),
    ("Confirmar case de sucesso com cliente (VALE / Klabin / Aegea)", "Fran / Sócios", "30/06/2026"),
    ("Confirmar speakers e apresentadores internos", "Sócios", "30/06/2026"),
    ("Confirmar presença do time (Sócios + Fran)", "Fran", "30/06/2026"),
    ("Confirmar participação de especialistas SAP para demo", "Fran", "30/06/2026"),
    ("Montar lista de convidados (clientes + prospects)", "Fran", "15/07/2026"),
    ("Segmentar lista por perfil (MQL, prospect, cliente ativo)", "Fran", "15/07/2026"),
]
for i, t in enumerate(items_plan):
    row = item(ws1, row, *t, zebra=(i%2==0))

# ---- PRODUÇÃO DE MATERIAIS ----
row = section_header(ws1, row, "🎨  PRODUÇÃO DE MATERIAIS")
items_mat = [
    ("Criar landing page / formulário de inscrição", "Fran / Design", "14/07/2026"),
    ("Produzir card WhatsApp do evento", "Fran / Design", "14/07/2026"),
    ("Redigir e-mail 01 — 1º convite", "Fran", "18/07/2026"),
    ("Redigir e-mail 02 — 2º convite", "Fran", "25/07/2026"),
    ("Redigir e-mail 03 — 3º convite", "Fran", "01/08/2026"),
    ("Redigir e-mail 04 — 4º convite", "Fran", "08/08/2026"),
    ("Redigir e-mail 05 — RSVP / últimas vagas", "Fran", "15/08/2026"),
    ("Redigir lembrete D-2", "Fran", "22/08/2026"),
    ("Redigir e-mail de agradecimento pós-evento", "Fran", "25/08/2026"),
    ("Criar anúncios LinkedIn — formato 1200x1200 (mín. 4 versões)", "Fran / Design", "18/07/2026"),
    ("Criar anúncios LinkedIn — formato 1200x627 (mín. 4 versões)", "Fran / Design", "18/07/2026"),
    ("Preparar apresentação abertura + BDC", "Sócios / Consultores", "18/08/2026"),
    ("Preparar slides do case de sucesso", "Cliente + Solveplan", "01/08/2026"),
    ("Criar QR Code da pesquisa de satisfação (Qualtrics)", "Fran", "20/08/2026"),
    ("Produzir posts pré-evento (LinkedIn + Instagram)", "Fran / Design", "18/07/2026"),
    ("Produzir posts pós-evento", "Fran / Design", "25/08/2026"),
    ("Preparar lista de presença (digital + impressa backup)", "Fran", "22/08/2026"),
]
for i, t in enumerate(items_mat):
    row = item(ws1, row, *t, zebra=(i%2==0))

# ---- DIVULGAÇÃO ----
row = section_header(ws1, row, "📢  DIVULGAÇÃO")
items_div = [
    ("Disparar E-mail 01 — 1º convite", "Fran", "21/07/2026"),
    ("Ativar LinkedIn Ads — campanha de convite", "Fran", "21/07/2026"),
    ("Publicar post pré-evento LinkedIn", "Fran", "21/07/2026"),
    ("Disparar E-mail 02 — 2º convite", "Fran", "28/07/2026"),
    ("Disparar E-mail 03 — 3º convite", "Fran", "04/08/2026"),
    ("Disparo WhatsApp para base qualificada", "Fran", "30/07 – 04/08"),
    ("Disparar E-mail 04 — 4º convite", "Fran", "11/08/2026"),
    ("Período RSVP — confirmação de presença", "Fran", "11/08 – 22/08"),
    ("Disparar E-mail 05 — RSVP / últimas vagas", "Fran", "18/08/2026"),
    ("Publicar post de contagem regressiva", "Fran", "20/08/2026"),
    ("Disparar lembrete D-2", "Fran", "24/08/2026"),
]
for i, t in enumerate(items_div):
    row = item(ws1, row, *t, zebra=(i%2==0))

# ---- LOGÍSTICA ----
row = section_header(ws1, row, "🏢  LOGÍSTICA E ESTRUTURA")
items_log = [
    ("Contratar coffee break para 70 pax (~R$ 4.400,00)", "Fran", "30/06/2026"),
    ("Comprar brindes — Moleskine e Caneta (70 unidades)", "Fran", "15/07/2026"),
    ("Confirmar camisetas Insider para o time", "Fran", "01/08/2026"),
    ("Reservar equipamento — projetor, áudio, tela", "SAP Brasil / Fran", "01/08/2026"),
    ("Confirmar acesso à internet no auditório", "SAP Brasil / Fran", "01/08/2026"),
    ("Verificar disponibilidade da sala Experience para demo complementar", "Fran", "01/07/2026"),
    ("Visita técnica ao espaço", "Fran", "19/08/2026"),
    ("Montar plano de contingência (speaker, internet, coffee)", "Fran", "19/08/2026"),
]
for i, t in enumerate(items_log):
    row = item(ws1, row, *t, zebra=(i%2==0))

# ---- DIA DO EVENTO ----
row = section_header(ws1, row, "📅  DIA DO EVENTO — 26/08/2026")
items_dia = [
    ("Chegada 1h antes para montagem e setup", "Fran + Sócios", "26/08 — 07:30"),
    ("Testar projetor, áudio e internet", "Fran", "26/08 — 07:45"),
    ("Validar QR Code da pesquisa de satisfação", "Fran", "26/08 — 07:50"),
    ("Confirmar brindes disponíveis na recepção", "Fran", "26/08 — 08:00"),
    ("Confirmar lista de presença impressa (backup)", "Fran", "26/08 — 08:00"),
    ("Definir quem faz credenciamento", "Fran", "26/08 — 08:00"),
    ("Executar coffee break (08:30 – 09:00)", "Fran + fornecedor", "26/08 — 08:30"),
    ("Registrar presença dos participantes", "Fran", "26/08 — durante"),
    ("Capturar fotos e vídeos (Stories/Reels)", "A definir", "26/08 — durante"),
    ("Identificar e abordar leads quentes em tempo real", "Sócios", "26/08 — durante"),
    ("Registrar insights: dores, objeções, sistemas usados", "Fran + Sócios", "26/08 — após"),
    ("Distribuir brindes mediante QR Code da pesquisa", "Fran", "26/08 — durante/encerramento"),
]
for i, t in enumerate(items_dia):
    row = item(ws1, row, *t, zebra=(i%2==0))

# ---- PÓS-EVENTO ----
row = section_header(ws1, row, "📊  PÓS-EVENTO")
items_pos = [
    ("Disparar e-mail de agradecimento", "Fran", "27/08/2026"),
    ("Publicar post pós-evento LinkedIn + Instagram", "Fran", "27/08/2026"),
    ("Consolidar lista de presença com leads captados", "Fran", "29/08/2026"),
    ("Limpar e deduplicar dados de leads", "Fran", "29/08/2026"),
    ("Subir todos os contatos no HubSpot (origem: BDC Innovation Day 2026)", "Fran", "29/08/2026"),
    ("Classificar leads: Quente / Morno / Frio", "Fran + Sócios", "29/08/2026"),
    ("Criar tarefas de follow-up no HubSpot", "Fran + Comercial", "29/08/2026"),
    ("Follow-up D+1 — contato direto com leads quentes", "Comercial", "27/08/2026"),
    ("Follow-up D+3 — e-mail com material do evento", "Comercial", "29/08/2026"),
    ("Follow-up D+5 — ligação comercial para SQLs", "Comercial", "01/09/2026"),
    ("Follow-up D+7 — último contato / proposta de diagnóstico", "Comercial", "03/09/2026"),
    ("Gerar relatório de resultados (leads, pipeline, ROI)", "Fran", "05/09/2026"),
    ("Publicar Reels e conteúdo pós-evento", "Fran", "01-12/09/2026"),
    ("Registrar aprendizados para próxima edição", "Fran + Sócios", "05/09/2026"),
]
for i, t in enumerate(items_pos):
    row = item(ws1, row, *t, zebra=(i%2==0))

# ==========================================
# ABA 2 — CRONOGRAMA
# ==========================================
ws2 = wb.create_sheet("Cronograma")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14

ws2.merge_cells("A1:E1")
set_cell(ws2, 1, 1, "SAP BDC Innovation Day 2026 — Cronograma", bold=True, bg=AZUL, fg=BRANCO, h_align="center", size=14, border=False)
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:E2")
set_cell(ws2, 2, 1, "26 de agosto de 2026  |  SAP Brasil, São Paulo", bg=AMARELO, fg=AZUL, h_align="center", size=11, bold=True, border=False)

row = 4
for col, h in enumerate(["Fase / Período", "Atividade", "Responsável", "Data Limite", "Status"], 1):
    set_cell(ws2, row, col, h, bold=True, bg=AZUL, fg=BRANCO, h_align="center")
row += 1

fases = [
    ("JUN 2026 — Planejamento", [
        ("Confirmar espaço SAP Brasil", "Fran", "20/06/2026"),
        ("Definir nome do evento", "Fran", "20/06/2026"),
        ("Confirmar case e speakers", "Fran / Sócios", "30/06/2026"),
        ("Contratar coffee (70 pax)", "Fran", "30/06/2026"),
        ("Comprar brindes", "Fran", "15/07/2026"),
    ]),
    ("JUL 2026 — Criação", [
        ("Produzir materiais (landing page, e-mails, anúncios, card)", "Fran / Design", "14/07/2026"),
        ("Montar lista de convidados segmentada", "Fran", "15/07/2026"),
        ("E-mail 01 — 1º convite + LinkedIn Ads", "Fran", "21/07/2026"),
        ("E-mail 02 — 2º convite", "Fran", "28/07/2026"),
        ("Disparo WhatsApp", "Fran", "30/07 – 04/08"),
    ]),
    ("AGO 2026 — Divulgação e RSVP", [
        ("E-mail 03 — 3º convite", "Fran", "04/08/2026"),
        ("E-mail 04 — 4º convite", "Fran", "11/08/2026"),
        ("Período RSVP / confirmações", "Fran", "11/08 – 22/08"),
        ("E-mail 05 — RSVP / últimas vagas", "Fran", "18/08/2026"),
        ("Preparação final (materiais, brindes, lista)", "Fran", "25/08/2026"),
        ("Lembrete D-2", "Fran", "24/08/2026"),
    ]),
    ("26/08 — EVENTO", [
        ("Execução do evento", "Fran + Sócios", "26/08/2026"),
        ("Captação e qualificação de leads", "Fran + Sócios", "26/08/2026"),
        ("Cobertura foto/vídeo (Stories/Reels)", "A definir", "26/08/2026"),
    ]),
    ("SET 2026 — Pós-evento", [
        ("E-mail de agradecimento + post pós-evento", "Fran", "27/08/2026"),
        ("Consolidar leads no HubSpot", "Fran", "29/08/2026"),
        ("Follow-up comercial D+1 a D+7", "Comercial", "27/08 – 03/09"),
        ("Relatório de resultados", "Fran", "05/09/2026"),
    ]),
]

zebra = False
for fase_nome, atividades in fases:
    ws2.merge_cells(f"A{row}:E{row}")
    set_cell(ws2, row, 1, fase_nome, bold=True, bg=AMARELO, fg=AZUL, h_align="left")
    ws2.row_dimensions[row].height = 18
    row += 1
    for ativ, resp, prazo in atividades:
        bg = CINZA_CLARO if zebra else BRANCO
        set_cell(ws2, row, 1, "", bg=bg)
        set_cell(ws2, row, 2, ativ, bg=bg)
        set_cell(ws2, row, 3, resp, bg=bg, h_align="center")
        set_cell(ws2, row, 4, prazo, bg=bg, h_align="center")
        set_cell(ws2, row, 5, "A fazer", bg=bg, h_align="center")
        ws2.row_dimensions[row].height = 16
        row += 1
        zebra = not zebra

# ==========================================
# ABA 3 — SEQUÊNCIA DE E-MAILS
# ==========================================
ws3 = wb.create_sheet("Sequência E-mails")
ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 35
ws3.column_dimensions["E"].width = 16
ws3.column_dimensions["F"].width = 16

ws3.merge_cells("A1:F1")
set_cell(ws3, 1, 1, "Sequência de E-mails — SAP BDC Innovation Day 2026", bold=True, bg=AZUL, fg=BRANCO, h_align="center", size=14, border=False)
ws3.row_dimensions[1].height = 30

row = 3
for col, h in enumerate(["✓", "E-mail", "Data Envio", "Objetivo", "Tx. Abertura Meta", "Status"], 1):
    set_cell(ws3, row, col, h, bold=True, bg=AZUL, fg=BRANCO, h_align="center")
row += 1

emails = [
    ("E-mail 01 — 1º Convite", "21/07/2026", "Primeiro contato — convidar para o evento", ">20%"),
    ("E-mail 02 — 2º Convite", "28/07/2026", "Reforço — destacar speakers e case", ">15%"),
    ("E-mail 03 — 3º Convite", "04/08/2026", "Urgência leve — vagas se esgotando", ">18%"),
    ("E-mail 04 — 4º Convite", "11/08/2026", "Última chamada geral", ">14%"),
    ("E-mail 05 — RSVP", "18/08/2026", "Confirmação de presença / últimas vagas", ">20%"),
    ("Lembrete D-2", "24/08/2026", "Lembrete final para confirmados", ">30%"),
    ("Agradecimento pós-evento", "27/08/2026", "Agradecimento + material + próximos passos", ">40%"),
    ("Confirmação de inscrição", "Automático", "Confirmar cadastro e fornecer detalhes", ">40%"),
    ("Recebimento de inscrição", "Automático", "Notificação interna de novo inscrito", "—"),
]
for i, (nome, data, obj, meta) in enumerate(emails):
    bg = CINZA_CLARO if i % 2 == 0 else BRANCO
    set_cell(ws3, row, 1, "☐", bg=bg, h_align="center")
    set_cell(ws3, row, 2, nome, bg=bg, bold=True)
    set_cell(ws3, row, 3, data, bg=bg, h_align="center")
    set_cell(ws3, row, 4, obj, bg=bg)
    set_cell(ws3, row, 5, meta, bg=bg, h_align="center")
    set_cell(ws3, row, 6, "A fazer", bg=bg, h_align="center")
    ws3.row_dimensions[row].height = 18
    row += 1

# ==========================================
# ABA 4 — ORÇAMENTO
# ==========================================
ws4 = wb.create_sheet("Orçamento")
ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 38
ws4.column_dimensions["C"].width = 18
ws4.column_dimensions["D"].width = 18
ws4.column_dimensions["E"].width = 18
ws4.column_dimensions["F"].width = 25

ws4.merge_cells("A1:F1")
set_cell(ws4, 1, 1, "Orçamento — SAP BDC Innovation Day 2026", bold=True, bg=AZUL, fg=BRANCO, h_align="center", size=14, border=False)
ws4.row_dimensions[1].height = 30
ws4.merge_cells("A2:F2")
set_cell(ws4, 2, 1, "Budget total: R$ 6.000,00", bg=AMARELO, fg=AZUL, h_align="center", size=12, bold=True, border=False)

row = 4
for col, h in enumerate(["#", "Item", "Valor Previsto", "Valor Realizado", "Diferença", "Observações"], 1):
    set_cell(ws4, row, col, h, bold=True, bg=AZUL, fg=BRANCO, h_align="center")
row += 1

itens_orc = [
    ("Coffee break (70 pax)", "R$ 4.400,00", ""),
    ("Brindes — Moleskine + Caneta (70 un.)", "R$ 800,00", "Condicionado ao QR Code da pesquisa"),
    ("LinkedIn Ads — campanha de convite", "R$ 800,00", "Formato 1200x1200 performou melhor na ed. anterior"),
    ("Contingência (10%)", "R$ 0,00", "Reserva inclusa no coffee (10% a mais contratado)"),
]
for i, (nome, previsto, obs) in enumerate(itens_orc):
    bg = CINZA_CLARO if i % 2 == 0 else BRANCO
    set_cell(ws4, row, 1, i+1, bg=bg, h_align="center")
    set_cell(ws4, row, 2, nome, bg=bg)
    set_cell(ws4, row, 3, previsto, bg=bg, h_align="center")
    set_cell(ws4, row, 4, "—", bg=bg, h_align="center")
    set_cell(ws4, row, 5, "—", bg=bg, h_align="center")
    set_cell(ws4, row, 6, obs, bg=bg)
    ws4.row_dimensions[row].height = 18
    row += 1

# Total
set_cell(ws4, row, 1, "", bg=AZUL)
set_cell(ws4, row, 2, "TOTAL", bold=True, bg=AZUL, fg=BRANCO)
set_cell(ws4, row, 3, "R$ 6.000,00", bold=True, bg=AZUL, fg=AMARELO, h_align="center")
set_cell(ws4, row, 4, "—", bg=AZUL, fg=BRANCO, h_align="center")
set_cell(ws4, row, 5, "—", bg=AZUL, fg=BRANCO, h_align="center")
set_cell(ws4, row, 6, "", bg=AZUL)
ws4.row_dimensions[row].height = 20
row += 2

# Indicadores
set_cell(ws4, row, 1, "", border=False)
set_cell(ws4, row, 2, "Custo por inscrito esperado (120 inscrições):", bold=True, fg=AZUL, border=False)
set_cell(ws4, row, 3, "~R$ 50,00", fg=AZUL, h_align="center", border=False)
row += 1
set_cell(ws4, row, 1, "", border=False)
set_cell(ws4, row, 2, "Custo por participante esperado (70 presentes):", bold=True, fg=AZUL, border=False)
set_cell(ws4, row, 3, "~R$ 86,00", fg=AZUL, h_align="center", border=False)
row += 1
set_cell(ws4, row, 1, "", border=False)
set_cell(ws4, row, 2, "Custo por oportunidade esperado (6 deals):", bold=True, fg=AZUL, border=False)
set_cell(ws4, row, 3, "~R$ 1.000,00", fg=AZUL, h_align="center", border=False)
row += 2
set_cell(ws4, row, 1, "", border=False)
set_cell(ws4, row, 2, "Referência fev/2025: R$ 5.034,80 investidos → 5 deals → ROI 9.824%", italic=True, fg=CINZA_MEDIO, border=False)

# ==========================================
# ABA 5 — RELATÓRIO DE RESULTADOS
# ==========================================
ws5 = wb.create_sheet("Relatório de Resultados")
ws5.column_dimensions["A"].width = 32
ws5.column_dimensions["B"].width = 20
ws5.column_dimensions["C"].width = 20
ws5.column_dimensions["D"].width = 20
ws5.column_dimensions["E"].width = 25

ws5.merge_cells("A1:E1")
set_cell(ws5, 1, 1, "Relatório de Resultados — SAP BDC Innovation Day 2026", bold=True, bg=AZUL, fg=BRANCO, h_align="center", size=14, border=False)
ws5.row_dimensions[1].height = 30
ws5.merge_cells("A2:E2")
set_cell(ws5, 2, 1, "Preencher após o evento (26/08/2026)", bg=AMARELO, fg=AZUL, h_align="center", size=11, bold=True, border=False)

row = 4
for col, h in enumerate(["Métrica", "Baseline Fev/2025", "Meta Ago/2026", "Realizado", "Observações"], 1):
    set_cell(ws5, row, col, h, bold=True, bg=AZUL, fg=BRANCO, h_align="center")
row += 1

metricas = [
    ("Inscrições", "108", "120", ""),
    ("Participantes", "63", "70", ""),
    ("Não participaram", "46", "50", ""),
    ("Taxa de comparecimento", "58%", "≥ 58%", ""),
    ("MQLs", "60", "70", ""),
    ("SQLs", "6", "7", ""),
    ("Oportunidades (deals)", "5", "6", ""),
    ("Pipeline gerado", "~R$ 2,5M", "R$ 3.000.000", ""),
    ("Investimento total", "R$ 5.034,80", "R$ 6.000,00", ""),
    ("ROI", "9.824%", ">9.000%", ""),
    ("Custo por inscrito", "R$ 46,62", "~R$ 50,00", ""),
    ("Custo por participante", "R$ 79,92", "~R$ 86,00", ""),
    ("Custo por oportunidade", "R$ 1.006,96", "~R$ 1.000,00", ""),
    ("Tx. abertura e-mail média", "~19%", ">20%", ""),
]
for i, (m, baseline, meta, obs) in enumerate(metricas):
    bg = CINZA_CLARO if i % 2 == 0 else BRANCO
    set_cell(ws5, row, 1, m, bold=True, bg=bg, fg=AZUL)
    set_cell(ws5, row, 2, baseline, bg=bg, h_align="center", fg="555555")
    set_cell(ws5, row, 3, meta, bold=True, bg=bg, h_align="center", fg=AZUL)
    set_cell(ws5, row, 4, "—", bg=bg, h_align="center")
    set_cell(ws5, row, 5, obs, bg=bg)
    ws5.row_dimensions[row].height = 18
    row += 1

row += 2
ws5.merge_cells(f"A{row}:E{row}")
set_cell(ws5, row, 1, "Aprendizados e Próximos Passos", bold=True, bg=AMARELO, fg=AZUL)
row += 1
for label in ["O que funcionou:", "O que não funcionou:", "O que melhorar:", "Próximos passos:"]:
    set_cell(ws5, row, 1, label, bold=True, fg=AZUL)
    ws5.merge_cells(f"B{row}:E{row}")
    set_cell(ws5, row, 2, "", bg=CINZA_CLARO)
    ws5.row_dimensions[row].height = 30
    row += 1

# Salvar
output_path = r"c:\Users\franc\solveplan.com\Roberto Molina - Marketing\1. MKT Estrategy\3. Agentes de IA\ccos-ratos\eventos\sap-bdc-innovation-day-2026\SAP_BDC_Innovation_Day_2026_Checklist.xlsx"
wb.save(output_path)
print(f"Excel salvo: {output_path}")

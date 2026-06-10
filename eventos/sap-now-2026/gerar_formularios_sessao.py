import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AZUL_ESCURO = "0D1B3E"
AZUL_MEDIO  = "1A56A0"
CINZA_CAMPO = "F2F2F2"
AMARELO     = "FFF9C4"
BRANCO      = "FFFFFF"
VERDE       = "E8F5E9"

def fill(h): return PatternFill("solid", fgColor=h)
def borda():
    t = Side(style="thin", color="BBBBBB")
    return Border(left=t, right=t, top=t, bottom=t)
def esquerda(wrap=True): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
def centro(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

def criar_formulario(cliente, solucao, titulo, descritivo, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulário SAP NOW"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 18

    # ── Cabeçalho ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = "SAP NOW AI TOUR BRAZIL 2026 — Conteúdo Patrocinadores"
    ws["A1"].font = Font(bold=True, color=BRANCO, size=13)
    ws["A1"].fill = fill(AZUL_ESCURO)
    ws["A1"].alignment = centro()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Formulário de Sessão de Conteúdo — {cliente}"
    ws["A2"].font = Font(bold=True, color=BRANCO, size=11)
    ws["A2"].fill = fill(AZUL_MEDIO)
    ws["A2"].alignment = centro()
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:D3")
    ws["A3"] = "⚠  Prazo de preenchimento: 26/06/2026   |   Sessão: 20 min   |   Até 2 palestrantes   |   Auditório aberto 50 lugares"
    ws["A3"].font = Font(bold=True, color="C0392B", size=10)
    ws["A3"].fill = fill(AMARELO)
    ws["A3"].alignment = centro()
    ws.row_dimensions[3].height = 20

    # ── Helper para linha de campo ──────────────────────────────────────────────
    def campo(ws, row, num, label, valor="", obs="", obrigatorio=True, altura=30, bg=CINZA_CAMPO):
        asterisco = " *" if obrigatorio else ""
        ws[f"A{row}"] = num
        ws[f"A{row}"].font = Font(bold=True, color=BRANCO, size=9)
        ws[f"A{row}"].fill = fill(AZUL_ESCURO)
        ws[f"A{row}"].alignment = centro()
        ws[f"A{row}"].border = borda()

        ws[f"B{row}"] = label + asterisco
        ws[f"B{row}"].font = Font(bold=True, size=10)
        ws[f"B{row}"].fill = fill(AZUL_MEDIO if not valor else BRANCO)
        ws[f"B{row}"].alignment = esquerda()
        ws[f"B{row}"].border = borda()
        ws[f"B{row}"].font = Font(bold=True, color=BRANCO if not valor else "000000", size=10)
        ws[f"B{row}"].fill = fill(AZUL_MEDIO) if not valor else fill(BRANCO)

        ws[f"C{row}"] = valor if valor else "— preencher —"
        ws[f"C{row}"].font = Font(size=10, color="000000" if valor else "999999",
                                  bold=True if valor else False)
        ws[f"C{row}"].fill = fill(VERDE) if valor else fill(CINZA_CAMPO)
        ws[f"C{row}"].alignment = esquerda()
        ws[f"C{row}"].border = borda()

        ws[f"D{row}"] = obs
        ws[f"D{row}"].font = Font(size=9, color="666666", italic=True)
        ws[f"D{row}"].fill = fill(AMARELO) if obs else fill(BRANCO)
        ws[f"D{row}"].alignment = esquerda()
        ws[f"D{row}"].border = borda()

        ws.row_dimensions[row].height = altura

    def secao(ws, row, titulo_secao):
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = titulo_secao
        ws[f"A{row}"].font = Font(bold=True, color=BRANCO, size=10)
        ws[f"A{row}"].fill = fill(AZUL_ESCURO)
        ws[f"A{row}"].alignment = esquerda()
        ws.row_dimensions[row].height = 20

    r = 4

    # ── Seção: Responsável ──────────────────────────────────────────────────────
    secao(ws, r, "RESPONSÁVEL PELA SESSÃO"); r += 1
    campo(ws, r, 1,  "Nome Solicitante",         "Francielle Beline"); r += 1
    campo(ws, r, 2,  "Empresa",                  "Solveplan"); r += 1
    campo(ws, r, 3,  "E-mail corporativo",        "francielle.beline@solveplan.com"); r += 1
    campo(ws, r, 4,  "Telefone (WhatsApp)",       "", obs="Preencher"); r += 1
    campo(ws, r, 5,  "Cota de patrocínio",        "Gold"); r += 1

    # ── Seção: Conteúdo ─────────────────────────────────────────────────────────
    r += 1
    secao(ws, r, "CONTEÚDO DA PALESTRA"); r += 1
    campo(ws, r, 6,  "Título da Palestra",        titulo,
          obs=f"{len(titulo)} / 100 caracteres", altura=35); r += 1
    campo(ws, r, 7,  "Descritivo da Palestra",    descritivo,
          obs=f"{len(descritivo)} / 350 caracteres", altura=80); r += 1
    campo(ws, r, 8,  "Tema primário",             "Caso de Sucesso (Customer Success Story)"); r += 1
    campo(ws, r, 9,  "Tipo de caso de sucesso",   "Cliente será um dos palestrantes ou estará presente no evento",
          obs="Confirmar com o cliente"); r += 1

    # ── Seção: Palestrante 1 ────────────────────────────────────────────────────
    r += 1
    secao(ws, r, f"PALESTRANTE 1 — {cliente.upper()}"); r += 1
    campo(ws, r, 10, "Empresa do palestrante 1",  cliente); r += 1
    campo(ws, r, 11, "Nome do palestrante 1",     "", obs="Confirmar com o cliente"); r += 1
    campo(ws, r, 12, "Cargo do palestrante 1",    "", obs="Confirmar com o cliente"); r += 1
    campo(ws, r, 13, "E-mail corporativo",        "", obs="Confirmar com o cliente"); r += 1
    campo(ws, r, 14, "Celular palestrante 1",     "", obs="Confirmar com o cliente"); r += 1

    # ── Seção: Palestrante 2 ────────────────────────────────────────────────────
    r += 1
    secao(ws, r, "PALESTRANTE 2 — SOLVEPLAN"); r += 1
    campo(ws, r, 15, "Empresa do palestrante 2",  "Solveplan"); r += 1
    campo(ws, r, 16, "Nome do palestrante 2",     "", obs="Alexandre Kuntgen ou Andrey?"); r += 1
    campo(ws, r, 17, "Cargo do palestrante 2",    "", obs="Confirmar internamente"); r += 1
    campo(ws, r, 18, "E-mail corporativo",        "", obs="Confirmar internamente"); r += 1
    campo(ws, r, 19, "Celular palestrante 2",     "", obs="Confirmar internamente"); r += 1

    # ── Seção: Termo ────────────────────────────────────────────────────────────
    r += 1
    secao(ws, r, "TERMO DE CONSENTIMENTO"); r += 1
    campo(ws, r, 20, "Termo de consentimento",
          "Confirmo estar ciente que o preenchimento deste formulário não caracteriza imediata aprovação do conteúdo proposto.",
          obs="Marcar ao enviar", altura=40); r += 1

    # ── Rodapé ──────────────────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = f"Solução: {solucao}   |   Evento: SAP NOW AI Tour Brazil 2026   |   Estande 03   |   Cota Gold"
    ws[f"A{r}"].font = Font(italic=True, size=9, color="666666")
    ws[f"A{r}"].alignment = centro()
    ws.row_dimensions[r].height = 18

    ws.freeze_panes = "B5"
    wb.save(path)
    print(f"Salvo: {path}")


# ── KLABIN ─────────────────────────────────────────────────────────────────────
criar_formulario(
    cliente   = "Klabin",
    solucao   = "SAP Business Data Cloud (BDC)",
    titulo    = "Como a Klabin unificou dados e acelerou decisões com SAP Business Data Cloud",
    descritivo= "A Klabin, maior produtora de papel e embalagens do Brasil, enfrentava dados dispersos e processos manuais que travavam decisões. Com SAP Business Data Cloud, unificou suas fontes de dados, ganhou visibilidade em tempo real e transformou analytics em vantagem competitiva. Conheça a jornada e os resultados.",
    path      = r"c:\Users\franc\solveplan.com\Roberto Molina - Marketing\1. MKT Estrategy\3. Agentes de IA\ccos-ratos\eventos\sap-now-2026\sessao-klabin-sap-now-2026.xlsx"
)

# ── OUROFINO ───────────────────────────────────────────────────────────────────
criar_formulario(
    cliente   = "Ourofino Agro",
    solucao   = "SAP Business Data Cloud (BDC) + SAP Analytics Cloud Planning",
    titulo    = "Ourofino Agro: planejamento integrado e dados unificados com SAP BDC",
    descritivo= "A Ourofino Agro operava com planejamento financeiro fragmentado e dados desconectados entre áreas. Com SAP Business Data Cloud e SAP Analytics Cloud Planning, integrou orçamento, forecast e consolidação em um único ambiente. O resultado: ciclos de planejamento mais rápidos, maior acuracidade e decisões baseadas em dados reais.",
    path      = r"c:\Users\franc\solveplan.com\Roberto Molina - Marketing\1. MKT Estrategy\3. Agentes de IA\ccos-ratos\eventos\sap-now-2026\sessao-ourofino-sap-now-2026.xlsx"
)

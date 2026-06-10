import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Checklist SAP NOW 2026"

AZUL_ESCURO = "0D1B3E"
AZUL_HEADER = "1A56A0"
VERMELHO    = "C0392B"
BRANCO      = "FFFFFF"
VERDE       = "D4EDDA"
CINZA       = "F5F5F5"

def fill(h): return PatternFill("solid", fgColor=h)
def borda():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)
def centro(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def esquerda(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

for i, w in enumerate([55, 12, 14, 22, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("A1:E1")
ws["A1"] = "CHECKLIST SAP NOW AI TOUR 2026  —  09 e 10/set  |  Transamérica Expo Center  |  Estande 03"
ws["A1"].font = Font(bold=True, color=BRANCO, size=13)
ws["A1"].fill = fill(AZUL_ESCURO)
ws["A1"].alignment = centro()
ws.row_dimensions[1].height = 28

def secao(ws, row, titulo):
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = titulo
    c.font = Font(bold=True, color=BRANCO, size=11)
    c.fill = fill(VERMELHO)
    c.alignment = esquerda()
    ws.row_dimensions[row].height = 22

def cabecalho(ws, row):
    for col, h in enumerate(["Atividade", "Prazo", "Status", "Responsável", "Obs."], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color=BRANCO, size=10)
        c.fill = fill(AZUL_ESCURO)
        c.alignment = centro()
        c.border = borda()
    ws.row_dimensions[row].height = 18

def status_cor(s):
    s = str(s).upper()
    if "OK" in s:             return fill("C8E6C9")
    if "EM ANDAMENTO" in s:   return fill("FFF3CD")
    if "PENDENTE" in s:       return fill("FFEBEE")
    return fill(BRANCO)

def linhas(ws, row, dados):
    for i, (atv, prazo, status, resp, obs) in enumerate(dados):
        r = row + i
        bg = fill(CINZA) if i % 2 == 0 else fill(BRANCO)
        for col, val in enumerate([atv, prazo, status, resp, obs], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(size=10)
            c.border = borda()
            if col == 3:
                c.fill = status_cor(val)
                c.alignment = centro()
            elif col == 2:
                c.fill = bg
                c.alignment = centro()
            else:
                c.fill = bg
                c.alignment = esquerda()
        ws.row_dimensions[r].height = 28
    return row + len(dados)

cur = 2

# ══════════════════════════════════════════════════════════════════════════════
# PLANEJAMENTO
# ══════════════════════════════════════════════════════════════════════════════
secao(ws, cur, "PLANEJAMENTO"); cur += 1
cabecalho(ws, cur); cur += 1

planejamento = [
    # ── Concluídos ──
    ("Fechamento do contrato",                                                                    "31/dez", "Ok",           "Fran",    ""),
    ("Acesso ao manual do expositor",                                                             "20/mai", "Ok",           "Fran",    ""),
    ("Envio de logomarca para organizadora para site",                                            "29/mai", "Ok",           "Fran",    "Fundo branco, 1920x1080px. Formatos: PNG ou JPG"),
    ("Breve descritivo da empresa",                                                               "29/mai", "Ok",           "Fran",    "Texto até 400 caracteres + logotipo da empresa"),
    ("Envio de logomarca para impressão",                                                         "24/jun", "Ok",           "Fran",    "Vetor (.ai ou .eps, CMYK), fontes convertidas em curvas"),
    ("Solicitação para agência nossa logo para o site",                                           "29/mai", "Ok",           "Fran",    ""),
    ("Solicitação para agência nossa logo para a impressão",                                      "29/mai", "Ok",           "Fran",    ""),
    ("Solicitar 2 Coletores de dados",                                                            "15/jun", "Ok",           "Fran",    "R$ 577,72 (2x) = Solicitado em 10/06/26"),
    ("Cartão de visita",                                                                          "10/jun", "Ok",           "Fran",    ""),
    ("Aprovar orçamento de anúncios para SAP NOW 2026 (P/Hotéis)",                               "15/jul", "Ok",           "Fran",    ""),
    ("Solicitar assinatura do orçamento de anúncios para SAP NOW 2026 (P/Hotéis)",               "20/jun", "Ok",           "Fran",    ""),
    ("Criar cadência de reuniões com sócios",                                                     "05/set", "Ok",           "Fran",    ""),
    # ── Em andamento ──
    ("Definir sessão principal 1ª escolha: Klabin",                                              "30/mai", "Em andamento", "Sócios",  "Klabin"),
    ("Formulário de sessão principal — enviar para a SAP (1ª escolha: TBD)",                     "06/jun", "Em andamento", "",        "Klabin, enviar para a SAP"),
    # ── Pendentes ──
    ("Acompanhar formulário enviado para SAP",                                                    "30/jun", "Pendente",     "Fran",    ""),
    ("Prazo final para pagamento de serviços e equipamentos",                                     "30/jul", "Pendente",     "Fran",    ""),
    ("Aguardando NF para pagamento dos serviços e equipamentos",                                  "15/jul", "Pendente",     "Fran",    ""),
    ("Aprovação de arte Plaquinha estande",                                                       "07/ago", "Pendente",     "Fran",    ""),
    ("Envio dos pedidos de aprovação de ativações e brindes para SAP",                            "30/jul", "Pendente",     "Fran",    ""),
    ("Confirmação de data e horário da palestra",                                                 "01/ago", "Pendente",     "Fran",    "Sessão Klabin — data TBD"),
    ("Prazo final para envio das artes exclusivas de acordo com as cotas adquiridas",             "15/jul", "Pendente",     "Fran",    ""),
    ("Definir ativações e brindes",                                                               "22/jun", "Pendente",     "Fran",    ""),
    ("Prazo final para envio dos pedidos de aprovação de ativações e brindes",                    "31/jul", "Pendente",     "Fran",    "CRÍTICO — prazo fixo SAP"),
    ("Burocracias da participação (taxa de prefeitura)",                                           "",       "Pendente",     "Fran",    ""),
    ("Data final para envio dos conteúdos para a agenda",                                         "26/jun", "Pendente",     "Fran",    "CRÍTICO — prazo fixo SAP"),
    ("Data final de Inscrição de Credencial do patrocinador",                                     "28/ago", "Pendente",     "Fran",    ""),
    ("Envio de report com os inscritos da Solveplan",                                             "11/set", "Pendente",     "SAP",     ""),
    ("Retirada de coletor de dados — CAEX",                                                       "08/set", "Pendente",     "Fran",    "RedDoor — R$ 1.601,14"),
    ("Credenciais: Definir time que irá ao evento",                                               "15/ago", "Pendente",     "Fran",    "8 credenciais + 1 staff"),
    ("Vídeos para TV principal",                                                                  "01/ago", "Pendente",     "Fran",    "R$ 5.280,00 — SAP com LCA estará no evento"),
    ("Credenciamento de montagem",                                                                "04/set", "Pendente",     "Fran",    ""),
    ("Finalização apresentação do Cliente",                                                       "",       "Pendente",     "Fran",    ""),
    ("Enviar apresentação do Cliente para SAP",                                                   "",       "Pendente",     "Fran",    ""),
    ("Pedir hotel evento para time",                                                              "01/ago", "Pendente",     "Viviane", ""),
    ("Criar arte para display de balcão com QR Code",                                            "15/ago", "Pendente",     "Fran",    ""),
    ("Definir clientes a serem convidados (até 10)",                                              "",       "Pendente",     "Sócios",  "Klabin, Aço Cearense, Citrosuco, Eurofarma, Grupo Comporte, M. Dias Branco, Usina da Pedra, Cocal, Zaffari, MarcoPolo"),
    ("Planejar anúncios de geolocalização durante os dias do evento",                             "30/jun", "Pendente",     "Fran",    ""),
    ("Criar anúncios de geolocalização durante os dias do evento",                                "15/jul", "Pendente",     "Fran",    ""),
    ("Enviar assinatura do orçamento de anúncios para SAP NOW 2026 (P/Hotéis) p/ Eletromidia",  "30/jun", "Pendente",     "Fran",    ""),
    ("Vídeo do anúncio para SAP NOW 2026 (P/Hotéis)",                                            "15/jul", "Pendente",     "Fran",    ""),
    ("Retirada de credencial",                                                                    "08/set", "Pendente",     "Todos",   "CAEX 08h-14h ou 09/set das 7h-18h"),
    ("Ir na SVPL pegar materiais e brindes",                                                      "08/ago", "Pendente",     "Andrey",  ""),
    ("Adaptador de rede para PC",                                                                 "",       "Pendente",     "Fran",    ""),
    ("TV + suporte + cabos",                                                                      "",       "Pendente",     "Fran",    ""),
    ("1 Tablet e peças para demo",                                                                "",       "Pendente",     "Andrey",  ""),
    ("Layout do stand",                                                                           "",       "Pendente",     "Fran",    ""),
    ("Orçamento stand com a agência",                                                             "10/jun", "Pendente",     "Fran",    "R$ 0,00"),
    ("Pgto stand layout da agência",                                                              "",       "Pendente",     "Fran",    ""),
    ("Criação das artes do DOOH",                                                                 "",       "Pendente",     "Fran",    ""),
    ("Aprovação das artes do DOOH",                                                               "",       "Pendente",     "Fran",    ""),
    ("Envio das artes do stand",                                                                  "15/jul", "Pendente",     "Fran",    ""),
    ("Camisetas para equipe",                                                                     "10/jul", "Pendente",     "Fran",    "27 camisetas + 12 polos"),
    ("Recebimento camisetas",                                                                     "10/ago", "Pendente",     "Fran",    ""),
    ("Recebimento Brindes",                                                                       "",       "Pendente",     "Fran",    "Cubo Mágico A Cuber Brasil — R$10.407,50"),
    ("", "", "Pendente", "", ""),
    ("", "", "Pendente", "", ""),
]
cur = linhas(ws, cur, planejamento)
cur += 1

# ══════════════════════════════════════════════════════════════════════════════
# COMUNICAÇÃO
# ══════════════════════════════════════════════════════════════════════════════
secao(ws, cur, "COMUNICAÇÃO"); cur += 1
cabecalho(ws, cur); cur += 1

comunicacao = [
    ("Inicio dos convites",                                                         "",  "Pendente", "Fran",              ""),
    ("Enviar lista de convites para a SAP",                                         "",  "Pendente", "Fran",              "Prazo: 10/jul — 10 C-levels, máx. 2 por empresa"),
    ("Página do evento no site",                                                    "",  "Pendente", "Fran",              ""),
    ("Formulário para QR Code do stand",                                            "",  "Pendente", "Fran",              "HubSpot form"),
    ("Fazer card para WhatsApp para convite",                                       "",  "Pendente", "Fran",              ""),
    ("Reunião com AEs, sócios e líderes de kickoff",                                "",  "Pendente", "Fran",              ""),
    ("Criação de legendas para posts de divulgação geral",                          "",  "Pendente", "Fran",              "POST 01 Lançamento / POST 02 Estande / POST 03 Amanhã / POST 04 Dia / POST 06 Agradecimento"),
    ("Criação de criativos para posts",                                             "",  "Pendente", "Fran",              ""),
    ("Programação dos posts via HubSpot",                                           "",  "Pendente", "Fran",              ""),
    ("Cronograma e-mails marketing de convite contatos gerais",                     "",  "Pendente", "Fran",              ""),
    ("Criação dos e-mails mkt de divulgação",                                       "",  "Pendente", "Fran",              ""),
    ("Criação de lista de e-mails no HubSpot para divulgação (e-mail mkt)",         "",  "Pendente", "Fran",              ""),
    ("Enviar email de alinhamento aos envolvidos do evento com as ações",           "",  "Pendente", "Fran",              ""),
    ("Criação do E-mail mkt com agradecimento pós-evento",                         "",  "Pendente", "Fran",              "Envio: 10/set"),
    ("Abordagem de slots SAP por SDR (Lucas/Larissa)",                              "",  "Pendente", "Geração de demanda","Gancho: demonstração ao vivo, Estande 03, 9 e 10/set"),
    ("", "", "Pendente", "", ""),
    ("", "", "Pendente", "", ""),
]
cur = linhas(ws, cur, comunicacao)
cur += 1

# ══════════════════════════════════════════════════════════════════════════════
# PRÉ-EVENTO
# ══════════════════════════════════════════════════════════════════════════════
secao(ws, cur, "PRÉ-EVENTO"); cur += 1
cabecalho(ws, cur); cur += 1

pre_evento = [
    ("Treinamento do time (kick off interno)",                     "Set/1",  "Pendente", "Fran",   "Após aprovação final — abordagem, dor, registro, o que NÃO falar"),
    ("Ativar anúncios DOOH + Google/LinkedIn",                     "07/set", "Pendente", "Fran",   "Iniciar 1 dia antes do evento"),
    ("Montar estande (4 acessos)",                                 "08/set", "Pendente", "Equipe", "Retirar pulseira CAEX das 09h-14h"),
    ("Testar ativação / totem (offline + captura de lead)",        "08/set", "Pendente", "Fran",   "Plano B: notebook com demo"),
    ("Testar QR Code e formulário HubSpot",                        "08/set", "Pendente", "Fran",   "URL curta como backup"),
    ("Testar vídeo looping na TV (pen drive)",                     "08/set", "Pendente", "Fran",   "2 pen drives — 1 de reserva"),
    ("Briefar speaker Klabin (local, horário, slides)",            "08/set", "Pendente", "Fran",   "Confirmar quem apresenta"),
    ("Conferir brindes, camisetas e materiais no estande",         "08/set", "Pendente", "Fran",   ""),
    ("", "", "Pendente", "", ""),
    ("", "", "Pendente", "", ""),
]
cur = linhas(ws, cur, pre_evento)
cur += 1

# ══════════════════════════════════════════════════════════════════════════════
# PÓS-EVENTO
# ══════════════════════════════════════════════════════════════════════════════
secao(ws, cur, "PÓS-EVENTO"); cur += 1
cabecalho(ws, cur); cur += 1

pos_evento = [
    ("Definir processo para prospecção dos leads gerados no evento",  "",       "Pendente", "Geração de demanda", ""),
    ("Recebimento do mailing dos participantes do evento (coletores)", "11/set", "Pendente", "Fran",              ""),
    ("Download do mailing EventsOnSite (estande e palestra)",          "",       "Pendente", "Fran",              "Origens separadas por canal"),
    ("Consolidar leads no CRM — tag SAP-NOW-2026",                    "D+1",    "Pendente", "Fran",              ""),
    ("Enviar e-mail de agradecimento personalizado",                   "D+1",    "Pendente", "Fran",              "Segmentar por dor declarada"),
    ("E-mail com valor (case Klabin) para leads mornos",               "D+3",    "Pendente", "Fran",              ""),
    ("Prospecção dos participantes",                                   "",       "Pendente", "Geração de demanda", "SLA: 48h"),
    ("Ligação / WhatsApp SDR — leads sem resposta",                    "D+5",    "Pendente", "SDR",               ""),
    ("E-mail resultados com avaliação",                                "",       "Pendente", "Fran",              ""),
    ("Report final",                                                   "",       "Pendente", "Fran",              "Comparar com 2025"),
    ("", "", "Pendente", "", ""),
    ("", "", "Pendente", "", ""),
]
cur = linhas(ws, cur, pos_evento)

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:E{cur}"

# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — ORÇAMENTO
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Orçamento")
for i, w in enumerate([38, 22, 18, 12, 12, 42], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:F1")
ws2["A1"] = "ORÇAMENTO — SAP NOW AI TOUR 2026"
ws2["A1"].font = Font(bold=True, color=BRANCO, size=13)
ws2["A1"].fill = fill(AZUL_ESCURO)
ws2["A1"].alignment = centro()
ws2.row_dimensions[1].height = 28

for col, h in enumerate(["Item", "Fornecedor", "Valor R$", "N° NF", "CC", "OBS"], 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = Font(bold=True, color=BRANCO, size=10)
    c.fill = fill(AZUL_ESCURO)
    c.alignment = centro()
    c.border = borda()
ws2.row_dimensions[2].height = 18

orcamento = [
    ("Cota de patrocínio Gold",                    "RedDoor",         265000.00, "", "", ""),
    ("Cota extra de comunicação",                  "—",                    0.00, "", "", "A confirmar se aplica"),
    ("Coletor de dados (2 un.)",                   "RedDoor",           1601.14, "", "", "R$ 577,72/un + fee e impostos"),
    ("Brindes — Cubo Mágico (250 un.)",            "A Cuber Brasil",   10407.50, "", "", "R$ 41,63/un — UV direto 5,6cm — renan@cuberbrasil.com / (19) 99606-4010"),
    ("Brindes — Bloco notas cubo BL31 (250 un.)",  "A definir",         8767.50, "", "", "R$ 35,07/un — 95x95x95mm, sik screen, azul"),
    ("Camisetas e polos",                          "A contratar",          0.00, "", "", "27 camisetas + 12 polos — aguardando RH"),
    ("Anúncios DOOH (hotéis SP)",                  "Eletromidia",       4988.69, "", "", "28 telas, 07-11/set"),
    ("Vídeos (looping + cases)",                   "A contratar",       5280.00, "", "", "SAP com LCA estará no evento"),
    ("Gráfica — display QR Code",                  "Kalunga",              0.00, "", "", "A orçar"),
    ("Arte do estande (agência)",                  "A contratar",          0.00, "", "", "Trainel + Painel + Balcão — a orçar"),
    ("Ativação (totem holográfico)",               "A contratar",          0.00, "", "", "A orçar — aprovação SAP obrigatória até 31/jul"),
    ("LinkedIn Ads + Google Ads",                  "A contratar",          0.00, "", "", "Geolocalização 07-11/set — a orçar"),
    ("Hotel da equipe",                            "A contratar",          0.00, "", "", "08 e 09/set — a orçar"),
]

for i, row in enumerate(orcamento, 3):
    bg = fill("EBF5FB") if i % 2 == 0 else fill(BRANCO)
    for col, val in enumerate(row, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.font = Font(size=10)
        c.border = borda()
        if col == 3:
            c.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(val, float) and val > 0:
                c.number_format = 'R$ #,##0.00'
            c.fill = bg
        else:
            c.fill = bg
            c.alignment = esquerda()
    ws2.row_dimensions[i].height = 22

tr = len(orcamento) + 3
ws2.merge_cells(f"A{tr}:B{tr}")
ws2[f"A{tr}"] = "TOTAL PARCIAL (itens com valor preenchido)"
ws2[f"A{tr}"].font = Font(bold=True, color=BRANCO, size=10)
ws2[f"A{tr}"].fill = fill(AZUL_ESCURO)
ws2[f"A{tr}"].alignment = centro()
ws2[f"A{tr}"].border = borda()
ws2[f"C{tr}"] = f"=SUM(C3:C{tr-1})"
ws2[f"C{tr}"].font = Font(bold=True, color=BRANCO, size=10)
ws2[f"C{tr}"].fill = fill(AZUL_ESCURO)
ws2[f"C{tr}"].alignment = Alignment(horizontal="right", vertical="center")
ws2[f"C{tr}"].number_format = 'R$ #,##0.00'
ws2[f"C{tr}"].border = borda()
ws2.row_dimensions[tr].height = 22

# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — COTAÇÕES BRINDES
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Cotações Brindes")
for i, w in enumerate([38, 20, 8, 6, 14, 8, 16, 16, 48], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:I1")
ws3["A1"] = "COTAÇÕES — BRINDES (CUBO MÁGICO)"
ws3["A1"].font = Font(bold=True, color=BRANCO, size=13)
ws3["A1"].fill = fill(AZUL_ESCURO)
ws3["A1"].alignment = centro()
ws3.row_dimensions[1].height = 28

for col, h in enumerate(["Item", "Fornecedor", "Qtde", "Unid.", "Prazo", "Frete", "Custo Unit. (R$)", "Custo Total (R$)", "Observações"], 1):
    c = ws3.cell(row=2, column=col, value=h)
    c.font = Font(bold=True, color=BRANCO, size=10)
    c.fill = fill(AZUL_ESCURO)
    c.alignment = centro()
    c.border = borda()
ws3.row_dimensions[2].height = 18

cotacoes = [
    ("Cubo Mágico Adesivado/Cromia 6 lados (5,2x5,2cm)", "Comercial Sisters", 250, "un", "12-15 dias", "CIF", 19.10,  4775.00,  "Adesivo Cromia 6 Logos — contato@comercialsisters.com.br"),
    ("Cubo mágico personalizado (5,5x5,5cm)",            "Imprimus",          250, "un", "15 dias",    "—",   13.95,  3487.50,  "Gravação Offset — guilherme@imprimus.com.br / (19) 3245-1615"),
    ("Cubo Mágico Profissional 3x3x3 (5,6cm)  ✓ ESCOLHIDO", "A Cuber Brasil", 250, "un", "4 dias",    "—",   41.63, 10407.50,  "UV direto no material — renan@cuberbrasil.com / (19) 99606-4010"),
    ("Cubo mágico 5,5x5,5x5,5",                         "ESP Brindes",        250, "un", "15 dias",    "CIF", 18.40,  4600.00,  "Papel adesivo com brilho envernizado — contato@espbrindes.com.br"),
    ("BL31 - Bloco notas cubo papelão reciclado",        "A definir",          250, "un", "—",          "—",   35.07,  8767.50,  "95x95x95mm, sik screen 01 cor, azul. 5 blocos + sticky notes + porta canetas"),
    ("Cubo Mágico Profissional 3x3x3 (5,6cm)",          "Still Promotion",    250, "un", "A combinar", "CIF", 66.90, 16725.00,  "UV direto no material — fabio@stillpromotion.com.br / 11-3906-0300"),
]

for i, row in enumerate(cotacoes, 3):
    chosen = "A Cuber Brasil" in str(row[1])
    bg = fill("C8E6C9") if chosen else (fill("EBF5FB") if i % 2 == 0 else fill(BRANCO))
    for col, val in enumerate(row, 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.font = Font(bold=True, size=10) if chosen else Font(size=10)
        c.border = borda()
        c.fill = bg
        if col in (7, 8):
            c.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(val, float) and val > 0:
                c.number_format = 'R$ #,##0.00'
        elif col in (3, 4, 5, 6):
            c.alignment = centro()
        else:
            c.alignment = esquerda()
    ws3.row_dimensions[i].height = 30

path = r"c:\Users\franc\solveplan.com\Roberto Molina - Marketing\1. MKT Estrategy\3. Agentes de IA\ccos-ratos\eventos\sap-now-2026\checklist-sap-now-2026.xlsx"
wb.save(path)
print(f"Salvo: {path}")
